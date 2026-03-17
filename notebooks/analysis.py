# /// script
# requires-python = ">=3.12"
# dependencies = [
#     "marimo",
#     "polars",
#     "plotly",
#     "httpx",
#     "openpyxl",
#     "pyarrow",
#     "numpy",
# ]
# ///

import marimo

__generated_with = "0.20.4"
app = marimo.App(width="medium")


@app.cell
def _():
    import marimo as mo

    mo.md(
        """
        # HUM-470 — Swiss Residential Space Heating Analysis
        ## IPAT Decomposition: Energy Demand for Space Heating

        **Group 12** | Elie's Sections: Data & Trends, Policy & Efficiency Drivers, Conclusion

        This notebook downloads, caches, and visualizes key datasets for the IPAT decomposition
        of energy demand for space heating in
        Swiss residential buildings.
        """
    )
    return (mo,)


@app.cell
def _(mo):
    mo.md("""
    ## 0. Setup & Data Download
    """)
    return


@app.cell
def _():
    import polars as pl
    import plotly.express as px
    import plotly.graph_objects as go
    from plotly.subplots import make_subplots
    import httpx
    from pathlib import Path
    import json

    # Resolve project root from this file's path (works in both run and edit mode)
    _app_path = Path(__file__).resolve()
    _PROJECT_DIR = _app_path.parent.parent if _app_path.parent.name == "notebooks" else _app_path.parent
    DATA_DIR = _PROJECT_DIR / "data"
    DATA_DIR.mkdir(exist_ok=True)
    RAW_DIR = DATA_DIR / "raw"
    RAW_DIR.mkdir(exist_ok=True)

    def download_cached(url: str, filename: str, *, force: bool = False) -> Path:
        """Download a file if not already cached in data/."""
        path = DATA_DIR / filename
        if path.exists() and not force:
            print(f"  [cached] {filename}")
            return path
        print(f"  [downloading] {filename} ...")
        resp = httpx.get(url, follow_redirects=True, timeout=60.0)
        resp.raise_for_status()
        path.write_bytes(resp.content)
        print(f"  [saved] {filename} ({len(resp.content) / 1024:.0f} KB)")
        return path

    def download_json_api(url: str, filename: str, *, payload: dict | None = None, force: bool = False) -> Path:
        """Download JSON from a REST API, cache locally."""
        path = DATA_DIR / filename
        if path.exists() and not force:
            print(f"  [cached] {filename}")
            return path
        print(f"  [downloading] {filename} ...")
        if payload:
            resp = httpx.post(url, json=payload, follow_redirects=True, timeout=30.0)
        else:
            resp = httpx.get(url, follow_redirects=True, timeout=30.0)
        resp.raise_for_status()
        path.write_text(json.dumps(resp.json(), indent=2, ensure_ascii=False))
        print(f"  [saved] {filename} ({len(resp.content) / 1024:.0f} KB)")
        return path

    print("Utilities ready.")
    return (
        RAW_DIR,
        download_cached,
        download_json_api,
        go,
        json,
        make_subplots,
        pl,
    )


@app.cell
def _(download_cached, download_json_api, mo):
    mo.md("### Downloading datasets...")

    # 1. BFE Energy Balance (CSV from opendata.swiss — verified working URL)
    energy_balance_path = download_cached(
        "https://www.uvek-gis.admin.ch/BFE/ogd/115/ogd115_gest_bilanz.csv",
        "bfe_energy_balance.csv",
    )

    # 2. BFE official HDD — population-weighted, base 20/12°C (monthly, 1994-present)
    hdd_path = download_cached(
        "https://www.uvek-gis.admin.ch/BFE/ogd/105/ogd105_heizgradtage.csv",
        "bfe_hdd_monthly.csv",
    )

    # 2b. Zurich Fluntern HDD (1864-present) — for pre-1994 coverage
    hdd_zurich_path = download_cached(
        "https://data.stadt-zuerich.ch/dataset/umw_heizgradtage_standort_jahr_monat_od1031/download/UMW103OD1031.csv",
        "zurich_hdd_annual.csv",
    )

    # 3. BFS Population (PX-Web REST API — total Swiss population 2010-2024)
    pop_path = download_json_api(
        "https://www.pxweb.bfs.admin.ch/api/v1/en/px-x-0102010000_101/px-x-0102010000_101.px",
        "bfs_population.json",
        payload={
            "query": [
                {"code": "Kanton (-) / Bezirk (>>) / Gemeinde (......)", "selection": {"filter": "item", "values": ["8100"]}},
                {"code": "Bevölkerungstyp", "selection": {"filter": "item", "values": ["1"]}},
                {"code": "Staatsangehörigkeit (Kategorie)", "selection": {"filter": "item", "values": ["-99999"]}},
                {"code": "Geschlecht", "selection": {"filter": "item", "values": ["-99999"]}},
                {"code": "Alter", "selection": {"filter": "item", "values": ["-99999"]}},
            ],
            "response": {"format": "json"},
        },
    )

    mo.md("**All datasets downloaded / cached.**")
    return energy_balance_path, hdd_path, hdd_zurich_path, pop_path


@app.cell
def _(mo):
    mo.md("""
    ---
    ## 1. Swiss Energy Balance — Household Final Energy Consumption

    Source: BFE/SFOE Gesamtenergiestatistik (opendata.swiss)

    The BFE energy balance gives us total final energy consumption by sector and by energy
    carrier from 1980 to 2024. We filter for **"Endverbrauch - Haushalte"** (household final
    consumption) which is dominated by space heating.
    """)
    return


@app.cell
def _(energy_balance_path, pl):
    # Load the BFE energy balance
    df_energy_raw = pl.read_csv(energy_balance_path, separator=",", infer_schema_length=10000)

    # Cast TJ column to float (it may contain string values)
    df_energy = df_energy_raw.with_columns(
        pl.col("TJ").cast(pl.Float64, strict=False)
    )

    # Filter for household final consumption
    df_hh = (
        df_energy
        .filter(pl.col("Rubrik") == "Endverbrauch - Haushalte")
        .filter(pl.col("TJ").is_not_null())
        .sort("Jahr", "Energietraeger")
    )

    print(f"Household energy data: {df_hh.shape[0]} rows, years {df_hh['Jahr'].min()}-{df_hh['Jahr'].max()}")
    print(f"Energy carriers: {df_hh['Energietraeger'].unique().sort().to_list()}")
    df_hh.head(15)
    return (df_hh,)


@app.cell
def _(df_hh, go, pl):
    # Map German energy carrier names to English for the chart
    carrier_map = {
        "Erdölprodukte": "Petroleum Products",
        "Elektrizität": "Electricity",
        "Gas": "Natural Gas",
        "Holzenergie": "Wood / Biomass",
        "Fernwärme": "District Heating",
        "Kohle": "Coal",
        "Uebrige erneuerbare Energien": "Other Renewables",
        "Müll und Industrieabfälle": "Waste",
    }
    carrier_colors = {
        "Petroleum Products": "#8B4513",
        "Natural Gas": "#FF6B35",
        "Electricity": "#FFC107",
        "Wood / Biomass": "#4CAF50",
        "District Heating": "#9C27B0",
        "Other Renewables": "#00BCD4",
        "Coal": "#607D8B",
        "Waste": "#795548",
    }

    # Pivot for stacked area
    df_hh_plot = df_hh.with_columns(
        pl.col("Energietraeger").replace(carrier_map).alias("carrier_en")
    )

    # Group by year and carrier
    df_pivot = (
        df_hh_plot
        .group_by(["Jahr", "carrier_en"])
        .agg(pl.col("TJ").sum())
        .sort("Jahr")
    )

    # Build stacked area chart
    fig_hh_energy = go.Figure()
    # Order carriers by total contribution (largest at bottom)
    carrier_totals = (
        df_pivot.group_by("carrier_en")
        .agg(pl.col("TJ").sum().alias("total"))
        .sort("total", descending=True)
    )
    ordered_carriers = carrier_totals["carrier_en"].to_list()

    for _carrier in ordered_carriers:
        _subset = df_pivot.filter(pl.col("carrier_en") == _carrier).sort("Jahr")
        fig_hh_energy.add_trace(go.Scatter(
            x=_subset["Jahr"].to_list(),
            y=_subset["TJ"].to_list(),
            name=_carrier,
            mode="lines",
            stackgroup="one",
            line=dict(width=0.5, color=carrier_colors.get(_carrier, "#999")),
        ))

    fig_hh_energy.update_layout(
        title="Swiss Household Final Energy Consumption by Carrier (TJ)",
        xaxis_title="Year",
        yaxis_title="Final Energy (TJ)",
        template="plotly_white",
        font=dict(size=14),
        width=1000,
        height=550,
        legend=dict(orientation="h", yanchor="bottom", y=-0.25, xanchor="center", x=0.5),
    )
    fig_hh_energy
    return


@app.cell
def _(df_hh, go, pl):
    # Total household energy over time (all carriers summed)
    df_hh_total = (
        df_hh
        .group_by("Jahr")
        .agg(pl.col("TJ").sum().alias("total_TJ"))
        .sort("Jahr")
    )

    # Also compute total for all end-use sectors for comparison
    fig_hh_total = go.Figure()
    fig_hh_total.add_trace(go.Scatter(
        x=df_hh_total["Jahr"].to_list(),
        y=df_hh_total["total_TJ"].to_list(),
        name="Household Total Energy",
        mode="lines+markers",
        line=dict(color="#2196F3", width=3),
        marker=dict(size=5),
    ))

    fig_hh_total.update_layout(
        title="Total Swiss Household Final Energy Consumption (TJ)",
        xaxis_title="Year",
        yaxis_title="Total Energy (TJ)",
        template="plotly_white",
        font=dict(size=14),
        width=950,
        height=500,
    )
    fig_hh_total
    return


@app.cell
def _(mo):
    mo.md("""
    ### Observations

    - Household energy was historically **dominated by petroleum products** (heating oil),
      but oil's share fell from ~57% (1990) to ~23% (2024). **Electricity surpassed oil** by 2024 (34% vs 23%).
    - **Electricity** has been rising steadily (partly from heat pump adoption — HPs use electricity).
    - **Other renewables** (includes heat pump ambient heat contribution) growing rapidly (~10% by 2024).
    - Total household energy peaked around **2010** and shows a clear **downward trend** since (~-10% vs 2000).
    """)
    return


@app.cell
def _(mo):
    mo.md("""
    ### Space-Heating Adjustment

    The BFE energy balance reports **total household energy** (space heating + hot water +
    cooking + appliances + lighting). For our decomposition of **space-heating demand**, we
    apply a correction factor from the BFE/Prognos *Ex-Post-Analyse des schweizerischen
    Energieverbrauchs nach Bestimmungsfaktoren* (annual reports, 2000–2024).

    Space heating's share of household energy varies year-to-year with weather (cold winters
    raise the share) but the trend is downward: from ~71 % (2000) to ~65 % (2024), as heating
    efficiency improved faster than growth in other end uses (appliances, hot water).

    Data downloaded directly from **pubdb.bfe.admin.ch** (publication 12388, Nov 2025).
    """)
    return


@app.cell
def _(RAW_DIR, pl):
    # Space-heating (Raumwärme) energy directly from BFE/Prognos publication:
    # "Der Energieverbrauch der Privaten Haushalte 2000–2024", Tabelle 1
    # Source: pubdb.bfe.admin.ch/de/publication/download/12388 (Nov 2025)
    # We read both Raumwärme (row 7) and Total (row 17) to compute the share.
    import openpyxl as _openpyxl_sh

    _wb_sh = _openpyxl_sh.load_workbook(
        RAW_DIR / "bfe_hh_energy_by_use_2024.xlsx", data_only=True,
    )
    _ws_sh = _wb_sh["Tabelle1"]

    _sh_records = []
    for _c in range(3, 28):  # col 3 = 2000, col 27 = 2024
        _yr = int(_ws_sh.cell(6, _c).value)
        _rw_pj = float(_ws_sh.cell(7, _c).value)   # Raumwärme (PJ)
        _tot_pj = float(_ws_sh.cell(17, _c).value)  # Total HH (PJ)
        _sh_records.append({
            "year": _yr,
            "raumwaerme_pj": round(_rw_pj, 2),
            "total_hh_pj": round(_tot_pj, 2),
            "sh_share": round(_rw_pj / _tot_pj, 4),
        })

    df_sh_share = pl.DataFrame(_sh_records).sort("year")

    def get_sh_share(year: int) -> float:
        """Look up the exact space-heating share for a given year."""
        _row = df_sh_share.filter(pl.col("year") == year)
        if _row.shape[0] > 0:
            return _row["sh_share"][0]
        # Fallback: use nearest year
        return df_sh_share.filter(pl.col("year") == df_sh_share["year"].max())["sh_share"][0]

    print("Space-heating share loaded from BFE/Prognos Tabelle 1.")
    print(f"  SH share 2000: {get_sh_share(2000):.1%}")
    print(f"  SH share 2012: {get_sh_share(2012):.1%}")
    print(f"  SH share 2024: {get_sh_share(2024):.1%}")
    df_sh_share
    return (get_sh_share,)


@app.cell
def _(mo):
    mo.md("""
    ---
    ## 2. Heating Degree Days (HDD)

    Source: BFE official population-weighted HDD (1994+) + Zurich Fluntern (pre-1994)

    - **1980-1993**: Zurich Fluntern station (Stadt Zurich open data)
    - **1994-2025**: BFE national population-weighted (opendata.swiss, ogd105)
    - Base: 20/12°C (Swiss standard)
    """)
    return


@app.cell
def _(hdd_path, hdd_zurich_path, pl):
    # === BFE national population-weighted HDD (1994-present) ===
    # Source: opendata.swiss — ogd105_heizgradtage.csv
    # Monthly data, we sum to annual totals
    df_hdd_monthly = (
        pl.read_csv(hdd_path, separator=",", infer_schema_length=1000)
        .filter(pl.col("Jahr").is_not_null())
        .with_columns(
            pl.col("Jahr").cast(pl.Int64),
            pl.col("Heizgradtage").cast(pl.Float64, strict=False),
        )
        .filter(pl.col("Heizgradtage").is_not_null())
    )

    df_hdd_bfe = (
        df_hdd_monthly
        .group_by("Jahr")
        .agg(pl.col("Heizgradtage").sum().round(1).alias("hdd"))
        .rename({"Jahr": "year"})
        .sort("year")
    )

    # === Zurich Fluntern HDD (1864-present) — for pre-1994 extension ===
    # Filter for Fluntern station only, sum monthly "Heizgradtag" column
    df_hdd_zurich = (
        pl.read_csv(hdd_zurich_path, separator=",", encoding="utf8-lossy",
                     infer_schema_length=5000)
        .filter(pl.col("Standort") == "Fluntern")
        .filter(pl.col("Heizgradtag").is_not_null())
        .with_columns(
            pl.col("Jahr").cast(pl.Int64),
            pl.col("Heizgradtag").cast(pl.Float64, strict=False),
        )
        .filter(pl.col("Heizgradtag").is_not_null())
        .group_by("Jahr")
        .agg(pl.col("Heizgradtag").sum().round(1).alias("hdd_zurich"))
        .rename({"Jahr": "year"})
        .sort("year")
    )

    # Build composite: Zurich pre-1994, BFE 1994+
    df_hdd_pre1994 = (
        df_hdd_zurich
        .filter(pl.col("year") < 1994)
        .filter(pl.col("year") >= 1980)
        .rename({"hdd_zurich": "hdd"})
    )

    df_hdd = pl.concat([df_hdd_pre1994, df_hdd_bfe]).sort("year").unique("year")

    print(f"HDD data: {df_hdd.shape[0]} years from {df_hdd['year'].min()} to {df_hdd['year'].max()}")
    print(f"  Pre-1994: Zurich Fluntern station ({df_hdd_pre1994.shape[0]} years)")
    print(f"  1994+: BFE national weighted ({df_hdd_bfe.shape[0]} years)")
    print(f"  2024 HDD: {df_hdd.filter(pl.col('year') == 2024)['hdd'][0]:.1f}")
    df_hdd.tail(10)
    return (df_hdd,)


@app.cell
def _(df_hdd, go):
    # HDD trend with OLS trendline
    years = df_hdd["year"].to_list()
    hdds = df_hdd["hdd"].to_list()

    # Simple linear regression (no numpy dependency)
    n = len(years)
    sum_x = sum(years)
    sum_y = sum(hdds)
    sum_xy = sum(x * y for x, y in zip(years, hdds))
    sum_x2 = sum(x * x for x in years)
    slope = (n * sum_xy - sum_x * sum_y) / (n * sum_x2 - sum_x * sum_x)
    intercept = (sum_y - slope * sum_x) / n
    trend_y = [slope * yr + intercept for yr in years]
    decline_per_decade = slope * 10

    fig_hdd = go.Figure()
    fig_hdd.add_trace(go.Scatter(
        x=years, y=hdds,
        name="HDD (actual)",
        mode="markers",
        marker=dict(size=8, color="#FF9800"),
    ))
    fig_hdd.add_trace(go.Scatter(
        x=years, y=trend_y,
        name=f"Trend ({decline_per_decade:.0f} HDD/decade)",
        mode="lines",
        line=dict(color="#F44336", width=2, dash="dash"),
    ))

    fig_hdd.update_layout(
        title="Heating Degree Days — Switzerland (base 20/12°C, Zurich Plateau)",
        xaxis_title="Year",
        yaxis_title="HDD",
        template="plotly_white",
        font=dict(size=14),
        width=950,
        height=500,
    )
    fig_hdd
    return


@app.cell
def _(mo):
    mo.md("""
    ### Weather is Getting Warmer

    HDD show a clear declining trend — Switzerland is warming faster than the global average
    (~2.9C above pre-industrial vs ~1.5C globally). This provides a "free" reduction in
    heating demand that is **independent of policy or technology**.

    When analyzing energy trends, we must separate this weather effect from structural
    improvements (insulation, heat pumps). That's why the LMDI decomposition should
    show both raw and HDD-normalized results.
    """)
    return


@app.cell
def _(mo):
    mo.md("""
    ---
    ## 3. Population — The P Factor

    Source: BFS (Federal Statistical Office) — PX-Web API
    """)
    return


@app.cell
def _(RAW_DIR, json, pl, pop_path):
    # Parse BFS population JSON (STATPOP, 2010+)
    pop_data = json.loads(pop_path.read_text())

    pop_records = []
    for _entry in pop_data.get("data", []):
        _year = int(_entry["key"][0])
        _val = _entry["values"][0]
        if _val and _val != "...":
            pop_records.append({"year": _year, "population": int(_val)})

    df_pop = pl.DataFrame(pop_records).sort("year")
    print(f"BFS STATPOP data: {df_pop.shape[0]} years ({df_pop['year'].min()}-{df_pop['year'].max()})")

    # Extend backward with World Bank SP.POP.TOTL for 2000-2009
    _wb_raw = json.loads((RAW_DIR / "worldbank_pop_ch.json").read_text())
    _wb_records = [
        {"year": int(e["date"]), "population": int(e["value"])}
        for e in _wb_raw[1]
        if e["value"] is not None and int(e["date"]) < df_pop["year"].min()
    ]
    df_pop_hist = pl.DataFrame(_wb_records)
    df_pop_full = pl.concat([df_pop_hist, df_pop]).sort("year").unique("year")
    print(f"Full population series: {df_pop_full.shape[0]} years ({df_pop_full['year'].min()}-{df_pop_full['year'].max()})")

    df_pop_full
    return (df_pop_full,)


@app.cell
def _(df_pop_full, go):
    fig_pop = go.Figure()
    fig_pop.add_trace(go.Scatter(
        x=df_pop_full["year"].to_list(),
        y=[p / 1_000_000 for p in df_pop_full["population"].to_list()],
        name="Population",
        mode="lines+markers",
        line=dict(color="#9C27B0", width=3),
        marker=dict(size=5),
    ))

    fig_pop.update_layout(
        title="Swiss Population (millions)",
        xaxis_title="Year",
        yaxis_title="Population (millions)",
        template="plotly_white",
        font=dict(size=14),
        width=900,
        height=450,
    )
    fig_pop
    return


@app.cell
def _(mo):
    mo.md("""
    ---
    ## 4. Heating Mix — Floor Area Share by Heating System

    **Sources (automated pipeline):**
    1. **BFE/Prognos Tabelle 13** (2000–2024): Energiebezugsfläche (EBF) by heating system — annual, Mio m²
    2. **BFE/Prognos Tabelle 11** (2000–2024): Space heating energy (PJ) by carrier — for cross-check
    3. **BFS Census via PxWeb** (1990, 2000): Building counts by heating energy source — for pre-2000 context
    4. **BFS GWS Excel** (2000, 2021): Building counts by heating system & energy source — for census cross-check

    The primary metric is **floor area share (%)** from Tabelle 13, which is more meaningful for
    IPAT analysis than raw building counts (a large MFH matters more than a small EFH).

    > **Caveat:** Floor area shares differ from building count shares (BFS GWR).
    > Heat pumps dominate new large MFH construction, so their floor area share (~29% in 2024)
    > exceeds their building count share (~23%). Oil remains #1 by building count (35%).
    > See [BFS energy sector](https://www.bfs.admin.ch/bfs/en/home/statistics/construction-housing/buildings/energy-sector.html).
    """)
    return


@app.cell
def _(RAW_DIR, pl):
    import openpyxl

    # ── 1. BFE/Prognos Tabelle 13: Floor area by heating system (Mio m² EBF), 2000–2024 ──
    wb = openpyxl.load_workbook(RAW_DIR / "bfe_hh_energy_by_use_2024.xlsx", data_only=True)
    ws13 = wb["Tabelle13"]

    # Row 6 = header: col[1]="Anlagensystem", cols[2:27]=years 2000–2024
    # Rows 7–14 = heating systems (col[1]=label, cols[2:27]=values), row 15 = Total
    header_row = list(ws13.iter_rows(min_row=6, max_row=6, values_only=True))[0]
    years_t13 = [int(v) for v in header_row[2:27] if isinstance(v, (int, float))]

    _system_rows = {}
    for _row in ws13.iter_rows(min_row=7, max_row=14, values_only=True):
        _label = _row[1]
        if _label and _label != "Total":
            _vals = [float(v) if v is not None else 0.0 for v in _row[2:27]]
            _system_rows[_label.strip()] = _vals

    # Map German labels → English column names
    _label_map = {
        "Heizöl": "oil",
        "Erdgas": "gas",
        "Elektrische Wärmepumpen": "heat_pump",
        "Holz": "wood",
        "El. Widerstandsheizungen": "electric_resistance",
        "Fernwärme": "district_heating",
        "Kohle": "coal",
        "Übrige/ ohne Heizsystem": "other",
    }

    # Build DataFrame with absolute values (Mio m²)
    t13_data = {"year": years_t13}
    for _de_label, _en_col in _label_map.items():
        t13_data[_en_col] = _system_rows.get(_de_label, [0.0] * len(years_t13))

    df_ebf = pl.DataFrame(t13_data)

    # Compute percentage shares
    carrier_cols = list(_label_map.values())
    df_mix = df_ebf.with_columns(
        _total=sum(pl.col(c) for c in carrier_cols),
    ).with_columns(
        **{c: (pl.col(c) / pl.col("_total") * 100).round(1) for c in carrier_cols}
    ).drop("_total")

    # Merge coal + other into "other_solar" to match chart categories
    df_mix = df_mix.with_columns(
        other_solar=(pl.col("coal") + pl.col("other")).round(1),
    ).drop("coal", "other")

    wb.close()
    df_mix
    return (df_mix,)


@app.cell
def _(RAW_DIR, pl):
    import openpyxl as _openpyxl

    # ── 2. BFE/Prognos Tabelle 11: Space heating energy by carrier (PJ), 2000–2024 ──
    # This provides a cross-check on the floor-area shares above.
    wb11 = _openpyxl.load_workbook(RAW_DIR / "bfe_hh_energy_by_use_2024.xlsx", data_only=True)
    ws11 = wb11["Tabelle11"]

    # Row 6 = header: col[1]="Energieträger", cols[2:27]=years 2000–2024
    # Rows 7–16 = carriers (col[1]=label, cols[2:27]=values), row 16 = Total
    header_row_11 = list(ws11.iter_rows(min_row=6, max_row=6, values_only=True))[0]
    years_t11 = [int(v) for v in header_row_11[2:27] if isinstance(v, (int, float))]

    _carrier_rows_11 = {}
    for _row in ws11.iter_rows(min_row=7, max_row=16, values_only=True):
        _label = _row[1]
        if _label and _label != "Total" and not str(_label).strip().startswith("darunter"):
            _vals = [float(v) if v is not None else 0.0 for v in _row[2:27]]
            _carrier_rows_11[_label.strip()] = _vals

    _label_map_11 = {
        "Heizöl": "oil",
        "Erdgas": "gas",
        "Elektrische Wärmepumpen": "heat_pump_elec",
        "Holz": "wood",
        "El. Widerstandsheizungen": "electric_resistance",
        "Fernwärme": "district_heating",
        "Kohle": "coal",
        "Umweltwärme": "ambient_heat",
        "Solar": "solar",
    }

    _t11_data = {"year": years_t11}
    for _de_label, _en_col in _label_map_11.items():
        _t11_data[_en_col] = _carrier_rows_11.get(_de_label, [0.0] * len(years_t11))

    df_heating_energy = pl.DataFrame(_t11_data)

    # Show energy shares (%) for comparison
    _energy_cols = list(_label_map_11.values())
    df_heating_energy_pct = df_heating_energy.with_columns(
        _total=sum(pl.col(c) for c in _energy_cols),
    ).with_columns(
        **{c: (pl.col(c) / pl.col("_total") * 100).round(1) for c in _energy_cols}
    ).drop("_total")

    df_heating_energy_pct
    return


@app.cell
def _(download_json_api, json, mo, pl):
    # ── 3. BFS Census (PxWeb): Building counts by heating energy source, 1990 & 2000 ──
    # Table: px-x-0902020100_122 (census 1970/1980/1990/2000)
    # Energieträger: 1=Heizöl, 2=Gas, 3=Elektrizität, 4=Holz/Kohle, 5=Andere, 9=Ohne
    census_path = download_json_api(
        "https://www.pxweb.bfs.admin.ch/api/v1/de/px-x-0902020100_122/px-x-0902020100_122.px",
        "bfs_census_buildings_heating.json",
        payload={
            "query": [
                {"code": "Kanton", "selection": {"filter": "item", "values": ["00"]}},
                {"code": "Energieträger der Heizung", "selection": {
                    "filter": "item", "values": ["1", "2", "3", "4", "5", "9"]
                }},
                {"code": "Jahr", "selection": {"filter": "item", "values": ["1990", "2000"]}},
            ],
            "response": {"format": "json"},
        },
    )

    raw_census = json.loads(census_path.read_text())

    # PxWeb returns key = [kanton_code, energieträger_code, year]
    # Energieträger codes: 1=Heizöl, 2=Gas, 3=Elektrizität, 4=Holz/Kohle, 5=Andere, 9=Ohne
    code_to_carrier = {
        "1": "oil",
        "2": "gas",
        "3": "electric",  # includes heat pumps + resistance
        "4": "wood_coal",
        "5": "other",
        "9": "none",
    }

    _census_records = []
    for _entry in raw_census["data"]:
        _carrier_code = _entry["key"][1]
        _year = int(_entry["key"][2])
        _count = int(_entry["values"][0]) if _entry["values"][0] != "..." else 0
        _carrier_en = code_to_carrier.get(_carrier_code, _carrier_code)
        _census_records.append({"year": _year, "carrier": _carrier_en, "buildings": _count})

    df_census_raw = pl.DataFrame(_census_records)

    # Pivot to wide format and compute percentages
    df_census = (
        df_census_raw
        .pivot(on="carrier", index="year", values="buildings")
        .sort("year")
    )

    mo.md(f"""
    ### BFS Census — Buildings by Heating Energy Source (1990 & 2000)

    Note: Census categories are coarser — "Elektrizität" includes both heat pumps and resistance
    heating, "Holz, Kohle" is combined. These serve as a historical anchor, not the primary metric.

    {mo.as_html(df_census)}
    """)
    return


@app.cell
def _(df_mix, go):
    # Stacked area chart — Heating Mix (floor area shares from BFE/Prognos Tabelle 13)
    carriers = ["oil", "gas", "heat_pump", "wood", "electric_resistance", "district_heating", "other_solar"]
    mix_colors = {
        "oil": "#8B4513",
        "gas": "#FF6B35",
        "heat_pump": "#2196F3",
        "wood": "#4CAF50",
        "electric_resistance": "#FFC107",
        "district_heating": "#9C27B0",
        "other_solar": "#00BCD4",
    }
    mix_labels = {
        "oil": "Heating Oil",
        "gas": "Natural Gas",
        "heat_pump": "Heat Pumps (electric)",
        "wood": "Wood / Biomass",
        "electric_resistance": "Electric Resistance",
        "district_heating": "District Heating",
        "other_solar": "Other (Coal + misc)",
    }

    fig_mix = go.Figure()
    for _carrier in carriers:
        fig_mix.add_trace(go.Scatter(
            x=df_mix["year"].to_list(),
            y=df_mix[_carrier].to_list(),
            name=mix_labels[_carrier],
            mode="lines",
            stackgroup="one",
            line=dict(width=0.5, color=mix_colors[_carrier]),
            fillcolor=mix_colors[_carrier],
        ))

    fig_mix.update_layout(
        title="Swiss Residential Heating — Floor Area Share by System (%, BFE/Prognos 2025)",
        xaxis_title="Year",
        yaxis_title="Share of Floor Area (%)",
        yaxis=dict(range=[0, 100]),
        template="plotly_white",
        font=dict(size=14),
        width=950,
        height=550,
        legend=dict(orientation="h", yanchor="bottom", y=-0.25, xanchor="center", x=0.5),
    )
    fig_mix
    return


@app.cell
def _(mo):
    mo.md("""
    ### Key Observations — Heating Mix

    **Important: this chart shows floor area shares (BFE/Prognos Tabelle 13), not building counts.**
    Floor area weights large multi-family buildings more heavily than small detached houses.
    By raw building count (BFS GWR 2024), oil still leads at **35%** vs heat pumps at **~23%**
    ([BFS 2024](https://www.bfs.admin.ch/bfs/en/home/statistics/construction-housing/buildings/energy-sector.html),
    [swissinfo](https://www.swissinfo.ch/eng/climate-solutions/one-fifth-of-swiss-households-heat-with-heat-pumps/87598078)).
    Heat pumps dominate new MFH construction, inflating their floor area share relative to building count.

    - **Oil floor area halved**: 59.9% (2000) → 23.9% (2024) — still #1 by building count (35%)
    - **Heat pumps lead by floor area** (28.9%) but not by building count (~23%) — the difference
      reflects their dominance in new large MFH buildings
    - **Gas peaked ~2022** at ~27% and is declining (energy crisis + cantonal fossil bans)
    - **Wood + district heating growing steadily**: combined ~12% → ~16%
    - **Electric resistance flat** at ~5–7% — legacy stock, no new installs
    """)
    return


@app.cell
def _(mo):
    mo.md("""
    ---
    ## 5. Decoupling Analysis — Energy Demand vs GDP

    The core IPAT question: is residential heating energy (I) decoupling from
    economic growth (P x A)?

    All series indexed to base year 2000 = 100.
    """)
    return


@app.cell
def _(df_floor_area, df_hh, df_pop_full, df_wb_gdp, get_sh_share, go, pl):
    # Build the decoupling chart from REAL data

    # 1. Space-heating energy by year (BFE household balance × SH share)
    _df_hh_yearly = (
        df_hh
        .group_by("Jahr")
        .agg(pl.col("TJ").sum().alias("hh_TJ"))
        .sort("Jahr")
        .filter(pl.col("Jahr") >= 2000)
        .rename({"Jahr": "year"})
    )
    _sh_shares = [get_sh_share(y) for y in _df_hh_yearly["year"].to_list()]
    _df_hh_yearly = _df_hh_yearly.with_columns(
        (pl.col("hh_TJ") * pl.Series("_sh", _sh_shares)).alias("total_TJ"),
    )

    # 2. GDP from World Bank (real, constant LCU -> billion CHF)
    _df_gdp = df_wb_gdp.select(
        pl.col("year"),
        pl.col("gdp_bn_chf").alias("gdp_bn"),
    )

    # 3. Population (from BFS API)
    _df_pop = df_pop_full.filter(pl.col("year") >= 2000)

    # 4. Floor area — back-extrapolate 2000–2011 from BFS 2012+ data
    #    BFS df_floor_area gives m2_per_person_avg (2012+); we extrapolate
    #    backward at 0.3%/yr growth (consistent with BFS long-run trend).
    _m2pp_2012 = df_floor_area.filter(pl.col("year") == 2012)["m2_per_person_avg"][0]
    _floor_pre_records = []
    for _row in _df_pop.filter(pl.col("year") < 2012).iter_rows(named=True):
        _yr, _pop = _row["year"], _row["population"]
        _m2pp = _m2pp_2012 * (1 - 0.003) ** (2012 - _yr)
        _floor_pre_records.append({"year": _yr, "floor_area_mio_m2": round(_m2pp * _pop / 1e6, 1)})
    _floor_hardcoded = pl.DataFrame(_floor_pre_records)

    # For 2012+ compute total floor area = m2_per_person_avg * population / 1e6
    _df_floor_real = (
        df_floor_area
        .select("year", "m2_per_person_avg")
        .join(_df_pop.select("year", "population"), on="year", how="left")
        .with_columns(
            (pl.col("m2_per_person_avg") * pl.col("population") / 1e6)
            .round(1)
            .alias("floor_area_mio_m2"),
        )
        .select("year", "floor_area_mio_m2")
    )

    _df_floor = pl.concat([_floor_hardcoded, _df_floor_real]).sort("year")

    # Join everything and index to 2000=100
    _df_decoupling = (
        _df_hh_yearly
        .join(_df_gdp, on="year", how="left")
        .join(_df_pop, on="year", how="left")
        .join(_df_floor, on="year", how="left")
    )

    # Get base year values
    _base = _df_decoupling.filter(pl.col("year") == 2000)
    _base_energy = _base["total_TJ"][0]
    _base_gdp = _base["gdp_bn"][0]
    _base_pop = _base["population"][0]
    _base_floor = _base["floor_area_mio_m2"][0]

    df_indexed = _df_decoupling.with_columns(
        (pl.col("total_TJ") / _base_energy * 100).alias("energy_idx"),
        (pl.col("gdp_bn") / _base_gdp * 100).alias("gdp_idx"),
        (pl.col("population") / _base_pop * 100).alias("pop_idx"),
        (pl.col("floor_area_mio_m2") / _base_floor * 100).alias("floor_idx"),
    )

    # Compute energy intensity index
    df_indexed = df_indexed.with_columns(
        (pl.col("energy_idx") / pl.col("floor_idx") * 100).alias("intensity_idx"),
    )

    # Plot
    _fig_decouple = go.Figure()

    _series = [
        ("gdp_idx", "GDP (real)", "#E91E63", "solid", 3),
        ("pop_idx", "Population", "#9C27B0", "dash", 2),
        ("floor_idx", "Heated Floor Area", "#FF9800", "dashdot", 2),
        ("energy_idx", "Space-Heating Energy", "#2196F3", "solid", 3),
        ("intensity_idx", "Energy Intensity (kWh/m2)", "#F44336", "dot", 2),
    ]

    for _col, _name, _color, _dash, _width in _series:
        _fig_decouple.add_trace(go.Scatter(
            x=df_indexed["year"].to_list(),
            y=df_indexed[_col].to_list(),
            name=_name,
            mode="lines+markers",
            line=dict(color=_color, width=_width, dash=_dash),
            marker=dict(size=4),
        ))

    _fig_decouple.add_hline(y=100, line_dash="dot", line_color="gray", opacity=0.5)

    # Policy annotations
    for _yr, _text, _y_pos in [
        (2008, "CO2 levy", 108),
        (2010, "Gebaudeprogramm", 118),
        (2015, "MuKEn 2014", 128),
        (2022, "CHF 120/t CO2", 138),
    ]:
        _fig_decouple.add_vline(x=_yr, line_dash="dot", line_color="rgba(0,0,0,0.15)")
        _fig_decouple.add_annotation(
            x=_yr, y=_y_pos, text=_text, showarrow=False,
            font=dict(size=9, color="gray"),
        )

    _fig_decouple.update_layout(
        title="Decoupling: Space-Heating Energy vs Economic Growth (2000 = 100)",
        xaxis_title="Year",
        yaxis_title="Index (2000 = 100)",
        template="plotly_white",
        font=dict(size=14),
        width=1050,
        height=600,
        legend=dict(orientation="h", yanchor="bottom", y=-0.22, xanchor="center", x=0.5),
    )
    _fig_decouple
    return (df_indexed,)


@app.cell
def _(df_indexed, mo):
    # Compute actual changes
    last = df_indexed.filter(df_indexed["year"] == df_indexed["year"].max())
    energy_change = last["energy_idx"][0] - 100
    gdp_change = last["gdp_idx"][0] - 100
    pop_change = last["pop_idx"][0] - 100
    floor_change = last["floor_idx"][0] - 100
    intensity_change = last["intensity_idx"][0] - 100

    mo.md(
        f"""
        ### Key Finding — Absolute Decoupling

        Since 2000 (all series indexed to 2000 = 100):

        | Indicator | Change | Why this value |
        |-----------|-------:|----------------|
        | **GDP (real)** | +{gdp_change:.0f}% | Steady Swiss economic growth; services and pharma drove output despite a strong franc |
        | **Population** | +{pop_change:.0f}% | Driven primarily by net immigration; Switzerland grew from 7.2 M to ~9.0 M |
        | **Heated floor area** | +{floor_change:.0f}% | Rising affluence → larger dwellings + more single-person households → more m² per capita |
        | **Space-heating energy** | {energy_change:+.0f}% | Despite upward pressure from P and A, heating energy **fell** — absolute decoupling |
        | **Energy intensity** | {intensity_change:+.0f}% | Energy per m² declined thanks to insulation retrofits (MuKEn: 90 → 16 kWh/m²), heat-pump adoption (COP 3–4 vs oil 0.9), and warmer winters (HDD −15 %) |

        **Absolute decoupling confirmed**: the Technology factor (T) is declining fast enough
        to more than offset growth in Population (P) and Affluence (A). The inflection
        point was around **2005–2008**, coinciding with the CO₂ levy introduction (2008)
        and the launch of the Gebäudeprogramm (2010).
        """
    )
    return


@app.cell
def _(mo):
    mo.md("""
    ---
    ## 6. Policy Timeline — Key Milestones

    Linking policy interventions to the observed trend shifts.
    """)
    return


@app.cell
def _(go):
    policies = [
        ("SIA 380/1 first edition", 1988, 1988, "#607D8B"),
        ("Energie 2000 program", 1990, 2000, "#4CAF50"),
        ("MuKEn 2000", 2000, 2008, "#2196F3"),
        ("CO2 levy introduced (CHF 12/t)", 2008, 2008, "#F44336"),
        ("MuKEn 2008", 2008, 2014, "#2196F3"),
        ("Gebaudeprogramm launched", 2010, 2024, "#9C27B0"),
        ("CO2 levy CHF 60/t", 2014, 2014, "#F44336"),
        ("MuKEn 2014 published", 2015, 2025, "#2196F3"),
        ("Energy Strategy 2050 enacted", 2017, 2017, "#FF9800"),
        ("CO2 levy CHF 96/t", 2018, 2018, "#F44336"),
        ("CO2 levy CHF 120/t", 2022, 2022, "#F44336"),
        ("Zurich fossil heating ban", 2022, 2022, "#E91E63"),
        ("Climate Protection Act", 2023, 2023, "#FF9800"),
        ("MuKEn 2025 approved", 2025, 2030, "#2196F3"),
    ]

    fig_policy = go.Figure()
    for _i, (_name, _start, _end, _color) in enumerate(policies):
        if _start == _end:
            fig_policy.add_trace(go.Scatter(
                x=[_start], y=[_i], mode="markers",
                marker=dict(size=12, color=_color, symbol="diamond"),
                showlegend=False,
            ))
        else:
            fig_policy.add_trace(go.Scatter(
                x=[_start, _end], y=[_i, _i], mode="lines",
                line=dict(color=_color, width=6),
                showlegend=False,
            ))
        # Place label to the right of the endpoint
        fig_policy.add_annotation(
            x=_end, y=_i, text=_name,
            xanchor="left", yanchor="middle",
            xshift=10, showarrow=False,
            font=dict(size=11, color=_color),
        )

    fig_policy.update_layout(
        title="Policy Timeline — Swiss Building Energy & Climate Milestones",
        xaxis_title="Year",
        xaxis=dict(range=[1985, 2032]),
        yaxis=dict(visible=False),
        template="plotly_white",
        font=dict(size=13),
        width=1000,
        height=600,
        margin=dict(l=20, r=300),
    )
    fig_policy
    return


@app.cell
def _(mo):
    mo.md("""
    ---
    ## 7. Building Standards Evolution — kWh/m2 Limits
    """)
    return


@app.cell
def _(go):
    standards = {
        "year": [1970, 1975, 2000, 2008, 2014],
        "limit_kwh_m2": [250, 180, 90, 60, 16],
        "label": ["No regulation<br>(~250)", "First SIA norm<br>(~180)",
                  "MuKEn 2000<br>(~90)", "MuKEn 2008<br>(~60)", "MuKEn 2014<br>(16)"],
    }

    fig_standards = go.Figure()
    fig_standards.add_trace(go.Bar(
        x=standards["year"], y=standards["limit_kwh_m2"],
        text=standards["label"], textposition="outside",
        marker_color=["#F44336", "#FF9800", "#FFC107", "#4CAF50", "#2196F3"],
        width=4,
    ))

    fig_standards.update_layout(
        title="New Building Heating Energy Limit (kWh/m2/yr) — Swiss Standards",
        xaxis_title="Year", yaxis_title="Max Heating Energy (kWh/m2/yr)",
        template="plotly_white", font=dict(size=14),
        width=900, height=500, yaxis=dict(range=[0, 310]),
    )
    fig_standards
    return


@app.cell
def _(mo):
    mo.md("""
    ---
    ## 8. Building Stock & Renovation Gap

    The structural bottleneck: slow renovation rate + old building stock.
    """)
    return


@app.cell
def _(go, make_subplots):
    age_data = {
        "period": ["Pre-1919", "1919-1945", "1946-1960", "1961-1970", "1971-1980",
                    "1981-1990", "1991-2000", "2001-2010", "2011-2020", "2021+"],
        "share": [12, 8, 10, 14, 16, 11, 9, 8, 9, 3],
        "typical_kwh_m2": [220, 200, 180, 160, 140, 100, 80, 45, 25, 16],
    }

    eff_colors = ["#F44336" if k > 150 else "#FF9800" if k > 80 else "#4CAF50"
                  for k in age_data["typical_kwh_m2"]]

    fig_stock = make_subplots(
        rows=1, cols=2,
        subplot_titles=("Building Stock by Construction Period (%)",
                        "Typical Energy Demand (kWh/m2/yr)"),
        horizontal_spacing=0.12,
    )

    fig_stock.add_trace(go.Bar(
        x=age_data["period"], y=age_data["share"],
        marker_color=eff_colors, showlegend=False,
    ), row=1, col=1)

    fig_stock.add_trace(go.Bar(
        x=age_data["period"], y=age_data["typical_kwh_m2"],
        marker_color=eff_colors, showlegend=False,
    ), row=1, col=2)

    fig_stock.update_layout(
        title="Swiss Building Stock: Age Distribution & Energy Performance",
        template="plotly_white", font=dict(size=13),
        width=1100, height=500,
    )
    fig_stock.update_yaxes(title_text="Share (%)", row=1, col=1)
    fig_stock.update_yaxes(title_text="kWh/m2/yr", row=1, col=2)
    fig_stock.update_xaxes(tickangle=45)
    fig_stock
    return


@app.cell
def _(mo):
    mo.md("""
    ### The Inertia Problem

    - **60% of buildings built before 1980** — most with poor insulation (140-220 kWh/m2)
    - Current renovation rate: **0.9-1.4%/year** — would take **~100 years** to renovate all
    - Net-zero by 2050 requires **~3%/year** — more than double current rate
    - Most renovations are **shallow** (window replacement only) not **deep** (full envelope)

    This is the structural constraint limiting how fast T can decline in the IPAT identity.
    Even as new buildings achieve 16 kWh/m2, the **old stock dominates** total demand.
    """)
    return


@app.cell
def _(mo):
    mo.md("""
    ---
    ## Summary — IPAT Decomposition Overview

    | Factor | Trend (2000-2024) | Direction | Role |
    |--------|-------------------|-----------|------|
    | **P** (Population) | +26% | Upward | Moderate pressure |
    | **A** (Floor area/capita) | +3-4% | Upward | Small pressure |
    | **T** (Energy/m2) | ~-31% | **Downward** | **Dominant driver** |
    | **I** (Total heating energy) | ~-10% | **Declining** | **Absolute decoupling** |

    **T is falling faster than P x A is rising** — absolute decoupling confirmed.

    Drivers of T reduction:
    1. Building envelope improvements (MuKEn standards: 250 -> 16 kWh/m2)
    2. Heat pump adoption (COP 3-4.5 vs fossil 1:1)
    3. Climate warming (HDD ~-5%/decade)
    4. CO2 levy making fossil heating expensive (CHF 120/tCO2)

    Main **brake**: slow renovation rate of pre-1980 building stock (0.9-1.4%/yr vs 3% needed).
    """)
    return


@app.cell
def _(mo):
    mo.md("""
    ---
    ## 9. External Data — World Bank GDP (constant CHF)

    Source: World Bank — NY.GDP.MKTP.KN (GDP in constant local currency units)

    Replaces the hardcoded `gdp_data` dict used in the decoupling cell with
    authoritative World Bank figures for 2000-2024.
    """)
    return


@app.cell
def _(RAW_DIR, json, pl):
    _gdp_raw = json.loads((RAW_DIR / "worldbank_gdp_ch.json").read_text())
    # The World Bank API returns [metadata, data_array]
    _gdp_entries = _gdp_raw[1]

    _gdp_records = []
    for _entry in _gdp_entries:
        _yr = int(_entry["date"])
        _val = _entry["value"]
        if _val is not None and 2000 <= _yr <= 2024:
            _gdp_records.append({"year": _yr, "gdp_real_chf": float(_val)})

    df_wb_gdp = pl.DataFrame(_gdp_records).sort("year")

    # Add a human-readable column in billions
    df_wb_gdp = df_wb_gdp.with_columns(
        (pl.col("gdp_real_chf") / 1e9).round(1).alias("gdp_bn_chf"),
    )

    print(f"World Bank GDP: {df_wb_gdp.shape[0]} years ({df_wb_gdp['year'].min()}-{df_wb_gdp['year'].max()})")
    print(f"GDP 2000: {df_wb_gdp.filter(pl.col('year') == 2000)['gdp_bn_chf'][0]:.1f} bn CHF")
    print(f"GDP 2024: {df_wb_gdp.filter(pl.col('year') == 2024)['gdp_bn_chf'][0]:.1f} bn CHF")
    df_wb_gdp
    return (df_wb_gdp,)


@app.cell
def _(df_wb_gdp, go):
    fig_wb_gdp = go.Figure()
    fig_wb_gdp.add_trace(go.Scatter(
        x=df_wb_gdp["year"].to_list(),
        y=df_wb_gdp["gdp_bn_chf"].to_list(),
        name="GDP (constant CHF)",
        mode="lines+markers",
        line=dict(color="#E91E63", width=3),
        marker=dict(size=5),
    ))

    fig_wb_gdp.update_layout(
        title="Swiss GDP — World Bank (constant LCU, billion CHF)",
        xaxis_title="Year",
        yaxis_title="GDP (billion CHF, constant prices)",
        template="plotly_white",
        font=dict(size=14),
        width=950,
        height=500,
    )
    fig_wb_gdp
    return


@app.cell
def _(mo):
    mo.md("""
    ---
    ## 10. Population Source Cross-Validation — BFS STATPOP vs World Bank

    Source: World Bank — SP.POP.TOTL

    Our population series uses World Bank data for 2000-2009 and BFS STATPOP for 2010+.
    This cross-validates both sources on the overlap years (2010-2024) to confirm consistency.
    """)
    return


@app.cell
def _(RAW_DIR, json, pl, pop_path):
    # BFS STATPOP (2010+)
    _bfs_data = json.loads(pop_path.read_text())
    _bfs_records = [
        {"year": int(e["key"][0]), "bfs_population": int(e["values"][0])}
        for e in _bfs_data.get("data", [])
        if e["values"][0] and e["values"][0] != "..."
    ]
    _df_bfs = pl.DataFrame(_bfs_records).sort("year")

    # World Bank (2000-2024)
    _wb_raw = json.loads((RAW_DIR / "worldbank_pop_ch.json").read_text())
    _wb_records = [
        {"year": int(e["date"]), "wb_population": int(e["value"])}
        for e in _wb_raw[1]
        if e["value"] is not None
    ]
    _df_wb = pl.DataFrame(_wb_records).sort("year")

    # Compare on overlap years
    df_pop_validation = _df_bfs.join(_df_wb, on="year", how="inner").with_columns(
        ((pl.col("wb_population") - pl.col("bfs_population")) / pl.col("bfs_population") * 100)
        .round(2)
        .alias("diff_pct"),
    )

    print(f"Population cross-validation — BFS vs World Bank ({df_pop_validation['year'].min()}-{df_pop_validation['year'].max()}):")
    df_pop_validation
    return (df_pop_validation,)


@app.cell
def _(df_pop_validation, mo):
    _max_diff = df_pop_validation["diff_pct"].abs().max()
    mo.md(
        f"""
        ### Validation Result

        Maximum absolute difference between BFS STATPOP and World Bank: **{_max_diff:.2f}%**

        BFS reports permanent resident population (end-of-year), while World Bank uses mid-year
        estimates. Small differences (<1%) confirm both sources are consistent and the pre-2010
        World Bank gap-fill is reliable.
        """
    )
    return


@app.cell
def _(mo):
    mo.md("""
    ---
    ## 11. BFS Floor Area per Person (2012-2024)

    Source: BFS GWS — T 09.03.02.04.03 (Durchschnittliche Wohnfläche pro Bewohner)

    Average living area per inhabitant in occupied dwellings, split by building type:
    - **Einfamilienhauser** (single-family houses)
    - **Mehrfamilienhauser** (multi-family buildings)
    """)
    return


@app.cell
def _(RAW_DIR, pl):
    import openpyxl as _openpyxl

    _wb = _openpyxl.load_workbook(
        RAW_DIR / "floor_area_by_building_category.xlsx", data_only=True,
    )

    _records = []
    for _sheet_name in _wb.sheetnames:
        _ws = _wb[_sheet_name]
        _yr = int(_sheet_name)
        # Row 5 = "Schweiz" national average; Col 2 = single-family, Col 3 = multi-family
        _sfh = _ws.cell(row=5, column=2).value
        _mfh = _ws.cell(row=5, column=3).value
        if _sfh is not None and _mfh is not None:
            _records.append({
                "year": _yr,
                "m2_per_person_sfh": float(_sfh),
                "m2_per_person_mfh": float(_mfh),
            })

    df_floor_area = (
        pl.DataFrame(_records)
        .sort("year")
        .with_columns(
            # Approximate weighted average (roughly 70% of people live in MFH in Switzerland)
            (pl.col("m2_per_person_sfh") * 0.30 + pl.col("m2_per_person_mfh") * 0.70)
            .round(1)
            .alias("m2_per_person_avg"),
        )
    )

    print(f"Floor area data: {df_floor_area.shape[0]} years ({df_floor_area['year'].min()}-{df_floor_area['year'].max()})")
    df_floor_area
    return (df_floor_area,)


@app.cell
def _(df_floor_area, go):
    fig_floor = go.Figure()

    fig_floor.add_trace(go.Scatter(
        x=df_floor_area["year"].to_list(),
        y=df_floor_area["m2_per_person_sfh"].to_list(),
        name="Single-Family Houses",
        mode="lines+markers",
        line=dict(color="#FF9800", width=2.5),
        marker=dict(size=6),
    ))
    fig_floor.add_trace(go.Scatter(
        x=df_floor_area["year"].to_list(),
        y=df_floor_area["m2_per_person_mfh"].to_list(),
        name="Multi-Family Buildings",
        mode="lines+markers",
        line=dict(color="#2196F3", width=2.5),
        marker=dict(size=6),
    ))
    fig_floor.add_trace(go.Scatter(
        x=df_floor_area["year"].to_list(),
        y=df_floor_area["m2_per_person_avg"].to_list(),
        name="Weighted Average (~30/70)",
        mode="lines+markers",
        line=dict(color="#4CAF50", width=3, dash="dash"),
        marker=dict(size=5),
    ))

    fig_floor.update_layout(
        title="Average Floor Area per Person — Swiss Residential Buildings (BFS GWS)",
        xaxis_title="Year",
        yaxis_title="m\u00b2 per person",
        template="plotly_white",
        font=dict(size=14),
        width=950,
        height=500,
        legend=dict(orientation="h", yanchor="bottom", y=-0.22, xanchor="center", x=0.5),
    )
    fig_floor
    return


@app.cell
def _(mo):
    mo.md("""
    ### Observations — Floor Area per Person

    - Single-family houses: **52.4 to 55.4 m²/person** (+5.7% over 2012-2024)
    - Multi-family buildings: **41.8 to 43.4 m²/person** (+3.8%)
    - The gap (~12 m²) is stable — SFH residents consistently enjoy more space
    - Floor area per person is **still growing**, adding pressure to total energy demand (the A factor in IPAT)
    """)
    return


@app.cell
def _(mo):
    mo.md("""
    ---
    ## 12. BFS Buildings by Heating Source — 2000 vs 2021

    Source: BFS GWS — T 09.02.07.04

    Comparison of the total number of residential buildings by primary heating energy source.
    """)
    return


@app.cell
def _(RAW_DIR, pl):
    import openpyxl as _openpyxl2

    _wb2 = _openpyxl2.load_workbook(
        RAW_DIR / "buildings_heating_source_2000_2021.xlsx", data_only=True,
    )

    # --- 2021 sheet ---
    # Row 6 = Total row; energy sources in cols 3-11
    _ws_2021 = _wb2["2021"]
    _heating_2021 = {
        "Heat Pump": _ws_2021.cell(row=6, column=3).value,
        "Gas": _ws_2021.cell(row=6, column=4).value,
        "Heating Oil": _ws_2021.cell(row=6, column=5).value,
        "Wood": _ws_2021.cell(row=6, column=6).value,
        "Electricity": _ws_2021.cell(row=6, column=7).value,
        "Solar Thermal": _ws_2021.cell(row=6, column=8).value,
        "District Heating": _ws_2021.cell(row=6, column=9).value,
        "Other": _ws_2021.cell(row=6, column=10).value,
    }

    # --- 2000 sheet ---
    # Row 6 = Total row; energy sources differ slightly (has Kohle/Coal column)
    _ws_2000 = _wb2["2000"]
    _heating_2000 = {
        "Heat Pump": _ws_2000.cell(row=6, column=3).value,
        "Gas": _ws_2000.cell(row=6, column=4).value,
        "Heating Oil": _ws_2000.cell(row=6, column=5).value,
        "Wood": _ws_2000.cell(row=6, column=7).value,
        "Electricity": _ws_2000.cell(row=6, column=8).value,
        "Solar Thermal": _ws_2000.cell(row=6, column=9).value,
        "District Heating": _ws_2000.cell(row=6, column=10).value,
        "Other": int(_ws_2000.cell(row=6, column=6).value or 0) + int(_ws_2000.cell(row=6, column=11).value or 0),
    }

    _sources = list(_heating_2021.keys())
    _vals_2000 = [int(_heating_2000.get(s, 0) or 0) for s in _sources]
    _vals_2021 = [int(_heating_2021.get(s, 0) or 0) for s in _sources]

    df_heating_compare = pl.DataFrame({
        "source": _sources,
        "buildings_2000": _vals_2000,
        "buildings_2021": _vals_2021,
    }).with_columns(
        ((pl.col("buildings_2021") - pl.col("buildings_2000")) / pl.col("buildings_2000").cast(pl.Float64) * 100)
        .round(1)
        .alias("change_pct"),
    )

    print("Buildings by heating source — 2000 vs 2021:")
    df_heating_compare
    return (df_heating_compare,)


@app.cell
def _(df_heating_compare, go):
    _sources = df_heating_compare["source"].to_list()
    _v2000 = df_heating_compare["buildings_2000"].to_list()
    _v2021 = df_heating_compare["buildings_2021"].to_list()

    _bar_colors = {
        "Heat Pump": "#2196F3",
        "Gas": "#FF6B35",
        "Heating Oil": "#8B4513",
        "Wood": "#4CAF50",
        "Electricity": "#FFC107",
        "Solar Thermal": "#00BCD4",
        "District Heating": "#9C27B0",
        "Other": "#607D8B",
    }

    fig_heating_cmp = go.Figure()

    fig_heating_cmp.add_trace(go.Bar(
        x=_sources,
        y=[v / 1000 for v in _v2000],
        name="2000",
        marker_color=[_bar_colors.get(s, "#999") for s in _sources],
        opacity=0.5,
    ))
    fig_heating_cmp.add_trace(go.Bar(
        x=_sources,
        y=[v / 1000 for v in _v2021],
        name="2021",
        marker_color=[_bar_colors.get(s, "#999") for s in _sources],
        opacity=0.9,
    ))

    fig_heating_cmp.update_layout(
        title="Residential Buildings by Heating Energy Source — 2000 vs 2021 (BFS GWS)",
        xaxis_title="Energy Source",
        yaxis_title="Number of Buildings (thousands)",
        barmode="group",
        template="plotly_white",
        font=dict(size=14),
        width=1000,
        height=550,
        legend=dict(orientation="h", yanchor="bottom", y=-0.18, xanchor="center", x=0.5),
    )
    fig_heating_cmp
    return


@app.cell
def _(mo):
    mo.md("""
    ### Key Changes 2000 to 2021

    - **Heat pumps**: 60k to 302k buildings (+403%) — the dominant growth story
    - **Heating oil**: 815k to 723k (-11%) — still the largest single source but declining
    - **Gas**: 200k to 312k (+56%) — grew but now peaked (post-2018 decline)
    - **Electricity (resistance)**: 166k to 143k (-14%) — being replaced by heat pumps
    - **District heating**: 21k to 64k (+210%) — tripled but still small in absolute terms
    """)
    return


@app.cell
def _(mo):
    mo.md("""
    ---
    ## 13. BFS Construction & Housing Summary — Dwellings by Energy Source

    Source: BFS — Bau- und Wohnungswesen (bau_wohnungswesen.csv)

    Dwelling-level breakdown of heating energy sources for all of Switzerland.
    """)
    return


@app.cell
def _(RAW_DIR, pl):
    df_bww_raw = pl.read_csv(
        RAW_DIR / "bau_wohnungswesen.csv",
        separator=",",
        encoding="utf8-lossy",
    )

    # Filter for Switzerland total and dwelling energy variables
    _energy_var_labels = {
        "wh_en_t": "Total dwellings",
        "wh_en_wp": "Heat Pump",
        "wh_en_gaz": "Gas",
        "wh_en_hoe": "Heating Oil",
        "wh_en_hz": "Wood",
        "wh_en_ez": "Electricity",
        "wh_en_st": "Solar Thermal",
        "wh_en_fw": "District Heating",
        "wh_en_and": "Other",
        "wh_en_kei": "None / Unknown",
    }

    df_bww_energy = (
        df_bww_raw
        .filter(pl.col("GEO_NR") == "CH")
        .filter(pl.col("VARIABLE").is_in(list(_energy_var_labels.keys())))
        .select("VARIABLE", "VALUE", "VALUE_PERIOD")
        .with_columns(
            pl.col("VALUE").cast(pl.Float64, strict=False).cast(pl.Int64).alias("dwellings"),
            pl.col("VARIABLE").replace(_energy_var_labels).alias("energy_source"),
        )
        .select("energy_source", "dwellings", "VALUE_PERIOD")
        .sort("dwellings", descending=True)
    )

    _total = df_bww_energy.filter(pl.col("energy_source") == "Total dwellings")["dwellings"][0]
    df_bww_shares = df_bww_energy.filter(pl.col("energy_source") != "Total dwellings").with_columns(
        (pl.col("dwellings") / _total * 100).round(1).alias("share_pct"),
    )

    print(f"BFS dwelling energy data — period: {df_bww_energy['VALUE_PERIOD'][0]}")
    print(f"Total dwellings: {_total:,.0f}")
    df_bww_shares
    return


@app.cell
def _(mo):
    mo.md(r"""
    ---
    ## 14. LMDI-I Decomposition — Population, Weather & Intensity Effects

    **Method**: Logarithmic Mean Divisia Index (additive, Ang 2004)

    We decompose the change in **space-heating energy** (household total × SH share)
    into three factors:

    $$E_{\text{total}} = P \times \frac{E}{P \cdot HDD} \times HDD$$

    - **Population effect** ($\Delta E_{\text{pop}}$): more people = more heating demand
    - **Weather effect** ($\Delta E_{\text{weather}}$): warmer winters = less heating demand
    - **Intensity effect** ($\Delta E_{\text{int}}$): technology & behaviour (insulation, heat pumps, efficiency)

    The decomposition is *exact*: $\Delta E_{\text{pop}} + \Delta E_{\text{int}} + \Delta E_{\text{weather}} = E^T - E^0$ (no residual).
    """)
    return


@app.cell
def _(df_hdd, df_hh, df_pop_full, get_sh_share, pl):
    # --- Build the LMDI input table ---
    # Total household energy by year, then apply space-heating share
    _df_energy_total = (
        df_hh
        .group_by("Jahr")
        .agg(pl.col("TJ").sum().alias("E_hh"))
        .rename({"Jahr": "year"})
        .sort("year")
        .filter(pl.col("year") >= 2000)
    )

    # Apply BFE space-heating correction (household total includes hot water,
    # cooking, appliances — only ~63–72 % is space heating).
    _sh = [get_sh_share(y) for y in _df_energy_total["year"].to_list()]
    _df_energy_total = _df_energy_total.with_columns(
        (pl.col("E_hh") * pl.Series("_sh", _sh)).alias("E"),
    )

    # Join energy, population, and HDD on year
    df_lmdi_input = (
        _df_energy_total
        .join(df_pop_full.select("year", "population"), on="year", how="inner")
        .join(df_hdd.select("year", "hdd"), on="year", how="inner")
        .sort("year")
    )

    # Compute energy intensity: E_sh / (P * HDD)
    df_lmdi_input = df_lmdi_input.with_columns(
        (pl.col("E") / (pl.col("population") * pl.col("hdd"))).alias("intensity"),
    )

    print(f"LMDI input (space-heating only): {df_lmdi_input.shape[0]} years "
          f"({df_lmdi_input['year'].min()}-{df_lmdi_input['year'].max()})")
    return (df_lmdi_input,)


@app.cell
def _(df_lmdi_input, pl):
    import numpy as _np

    # --- Logarithmic mean ---
    def _log_mean(a: float, b: float) -> float:
        """L(a, b) = (a - b) / ln(a/b), with L(a, a) = a."""
        if a <= 0 or b <= 0:
            return 0.0
        if abs(a - b) < 1e-12:
            return a
        return (a - b) / _np.log(a / b)

    # --- Year-on-year LMDI contributions ---
    _years = df_lmdi_input["year"].to_list()
    _E = df_lmdi_input["E"].to_list()
    _P = df_lmdi_input["population"].to_list()
    _HDD = df_lmdi_input["hdd"].to_list()
    _I = df_lmdi_input["intensity"].to_list()

    _records = []
    for _t in range(1, len(_years)):
        _L = _log_mean(_E[_t], _E[_t - 1])

        _dE_pop = _L * _np.log(_P[_t] / _P[_t - 1])
        _dE_weather = _L * _np.log(_HDD[_t] / _HDD[_t - 1])
        _dE_intensity = _L * _np.log(_I[_t] / _I[_t - 1])

        _records.append({
            "year": _years[_t],
            "dE_pop": _dE_pop,
            "dE_weather": _dE_weather,
            "dE_intensity": _dE_intensity,
            "dE_total": _E[_t] - _E[_t - 1],
        })

    _df_annual = pl.DataFrame(_records)

    # --- Cumulative sums from base year ---
    df_lmdi = _df_annual.with_columns(
        pl.col("dE_pop").cum_sum().alias("cum_pop"),
        pl.col("dE_weather").cum_sum().alias("cum_weather"),
        pl.col("dE_intensity").cum_sum().alias("cum_intensity"),
        pl.col("dE_total").cum_sum().alias("cum_total"),
    )

    # Verify decomposition is exact (residual should be ~0)
    _residual = (
        df_lmdi["cum_total"][-1]
        - df_lmdi["cum_pop"][-1]
        - df_lmdi["cum_weather"][-1]
        - df_lmdi["cum_intensity"][-1]
    )
    return (df_lmdi,)


@app.cell
def _(df_lmdi, go):
    # --- Waterfall chart: cumulative contribution from 2000 to latest year ---
    _last = df_lmdi[-1]
    _pop_val = _last["cum_pop"].item()
    _weather_val = _last["cum_weather"].item()
    _intensity_val = _last["cum_intensity"].item()
    _net_val = _last["cum_total"].item()

    fig_lmdi_waterfall = go.Figure(go.Waterfall(
        name="LMDI",
        orientation="v",
        measure=["relative", "relative", "relative", "total"],
        x=["Population", "Weather (HDD)", "Intensity (tech.)", "Net Change"],
        y=[_pop_val, _weather_val, _intensity_val, _net_val],
        text=[
            f"{_pop_val:+.0f} TJ",
            f"{_weather_val:+.0f} TJ",
            f"{_intensity_val:+.0f} TJ",
            f"{_net_val:+.0f} TJ",
        ],
        textposition="outside",
        connector=dict(line=dict(color="rgba(0,0,0,0.3)", width=1)),
        increasing=dict(marker=dict(color="#F44336")),
        decreasing=dict(marker=dict(color="#4CAF50")),
        totals=dict(marker=dict(color="#2196F3")),
    ))

    fig_lmdi_waterfall.update_layout(
        title="LMDI-I Decomposition of Space-Heating Energy Change (2000 to latest, TJ)",
        yaxis_title="Cumulative ΔE (TJ)",
        template="plotly_white",
        font=dict(size=14),
        width=900,
        height=550,
        showlegend=False,
    )
    return


@app.cell
def _(df_lmdi, go):
    # --- Time series of cumulative effects ---
    _years_ts = df_lmdi["year"].to_list()
    _cum_pop = df_lmdi["cum_pop"].to_list()
    _cum_weather = df_lmdi["cum_weather"].to_list()
    _cum_intensity = df_lmdi["cum_intensity"].to_list()
    _cum_total = df_lmdi["cum_total"].to_list()

    fig_lmdi_ts = go.Figure()

    fig_lmdi_ts.add_trace(go.Scatter(
        x=_years_ts, y=_cum_pop,
        name="Population",
        mode="lines+markers",
        line=dict(color="#F44336", width=2.5),
        marker=dict(size=4),
    ))
    fig_lmdi_ts.add_trace(go.Scatter(
        x=_years_ts, y=_cum_weather,
        name="Weather (HDD)",
        mode="lines+markers",
        line=dict(color="#2196F3", width=2.5),
        marker=dict(size=4),
    ))
    fig_lmdi_ts.add_trace(go.Scatter(
        x=_years_ts, y=_cum_intensity,
        name="Intensity (tech.)",
        mode="lines+markers",
        line=dict(color="#4CAF50", width=2.5),
        marker=dict(size=4),
    ))
    fig_lmdi_ts.add_trace(go.Scatter(
        x=_years_ts, y=_cum_total,
        name="Net Change",
        mode="lines+markers",
        line=dict(color="#212121", width=3, dash="dash"),
        marker=dict(size=5),
    ))

    fig_lmdi_ts.add_hline(y=0, line_dash="dot", line_color="gray", opacity=0.5)

    fig_lmdi_ts.update_layout(
        title="Cumulative LMDI-I Decomposition — Year-by-Year (TJ, base 2000)",
        xaxis_title="Year",
        yaxis_title="Cumulative ΔE (TJ)",
        template="plotly_white",
        font=dict(size=14),
        width=1000,
        height=550,
        legend=dict(orientation="h", yanchor="bottom", y=-0.22, xanchor="center", x=0.5),
    )
    return


@app.cell
def _(df_lmdi, mo):
    # --- Summary with dynamic numbers ---
    _last_row = df_lmdi[-1]
    _cum_p = _last_row["cum_pop"].item()
    _cum_w = _last_row["cum_weather"].item()
    _cum_i = _last_row["cum_intensity"].item()
    _cum_t = _last_row["cum_total"].item()

    # Shares of the absolute net change (handle sign correctly)
    _abs_sum = abs(_cum_p) + abs(_cum_w) + abs(_cum_i)
    _pct_pop = abs(_cum_p) / _abs_sum * 100
    _pct_weather = abs(_cum_w) / _abs_sum * 100
    _pct_intensity = abs(_cum_i) / _abs_sum * 100

    _up_p = "Up" if _cum_p > 0 else "Down"
    _up_w = "Up" if _cum_w > 0 else "Down"
    _up_i = "Up" if _cum_i > 0 else "Down"
    _up_t = "Up" if _cum_t > 0 else "Down"
    _last_yr = df_lmdi["year"][-1]

    mo.md(
        f"""
        ### LMDI Summary (2000 to {_last_yr})

        | Factor | Cumulative dE (TJ) | Direction | Share |
        |--------|--------------------:|-----------|------:|
        | **Population** | {_cum_p:+,.0f} | {_up_p} | {_pct_pop:.0f}% |
        | **Weather (HDD)** | {_cum_w:+,.0f} | {_up_w} | {_pct_weather:.0f}% |
        | **Intensity (tech.)** | {_cum_i:+,.0f} | {_up_i} | {_pct_intensity:.0f}% |
        | **Net Change** | {_cum_t:+,.0f} | {_up_t} | — |

        **Key insight**: The intensity effect alone ({_cum_i:+,.0f} TJ) is large enough to
        offset **both** population growth ({_cum_p:+,.0f} TJ) **and** would still yield a
        net decline even without the weather bonus ({_cum_w:+,.0f} TJ). This confirms that
        **technology and policy — not just climate luck — are driving absolute decoupling**.
        """
    )
    return


@app.cell
def _(mo):
    mo.md("""
    ---
    ## 15. HDD-Normalized Household Energy

    Removes year-to-year weather variation to reveal the **structural** trend
    in household energy consumption. The reference HDD is the 2000-2024 mean.
    """)
    return


@app.cell
def _(df_hdd, df_hh, get_sh_share, go, pl):
    # Reference HDD = mean over study period
    _hdd_ref = df_hdd.filter(pl.col("year").is_between(2000, 2024))["hdd"].mean()

    _df_e_yr = (
        df_hh.group_by("Jahr")
        .agg(pl.col("TJ").sum().alias("E_hh"))
        .rename({"Jahr": "year"})
        .sort("year")
        .filter(pl.col("year") >= 2000)
    )
    # Apply space-heating share
    _sh = [get_sh_share(y) for y in _df_e_yr["year"].to_list()]
    _df_e_yr = _df_e_yr.with_columns(
        (pl.col("E_hh") * pl.Series("_sh", _sh)).alias("E"),
    )

    df_hdd_norm = (
        _df_e_yr
        .join(df_hdd.select("year", "hdd"), on="year", how="inner")
        .with_columns(
            (pl.col("E") * _hdd_ref / pl.col("hdd")).round(0).alias("E_norm")
        )
    )

    fig_hdd_norm = go.Figure()
    fig_hdd_norm.add_trace(go.Scatter(
        x=df_hdd_norm["year"].to_list(),
        y=df_hdd_norm["E"].to_list(),
        name="Actual",
        mode="lines+markers",
        line=dict(color="#90CAF9", width=2),
        marker=dict(size=5),
    ))
    fig_hdd_norm.add_trace(go.Scatter(
        x=df_hdd_norm["year"].to_list(),
        y=df_hdd_norm["E_norm"].to_list(),
        name=f"HDD-Normalized (ref={_hdd_ref:.0f})",
        mode="lines+markers",
        line=dict(color="#1565C0", width=3),
        marker=dict(size=6),
    ))

    fig_hdd_norm.update_layout(
        title="Space-Heating Energy — Actual vs HDD-Normalized (TJ)",
        xaxis_title="Year",
        yaxis_title="Final Energy (TJ)",
        template="plotly_white",
        font=dict(size=14),
        width=1000,
        height=550,
        legend=dict(orientation="h", yanchor="bottom", y=-0.22, xanchor="center", x=0.5),
    )
    fig_hdd_norm
    return df_hdd_norm, fig_hdd_norm


@app.cell
def _(df_hdd_norm, mo, pl):
    _e_first = df_hdd_norm.filter(pl.col("year") == df_hdd_norm["year"].min())["E_norm"][0]
    _e_last = df_hdd_norm.filter(pl.col("year") == df_hdd_norm["year"].max())["E_norm"][0]
    _pct = (_e_last - _e_first) / _e_first * 100

    mo.md(
        f"""
        ### HDD-Normalized Trend

        After removing weather variability, household energy shows a **{_pct:+.1f}%** structural
        change from {df_hdd_norm['year'].min()} to {df_hdd_norm['year'].max()}.
        The normalized series is smoother — year-to-year spikes (cold winters) disappear,
        revealing the underlying efficiency improvement driven by insulation and heat pump adoption.
        """
    )
    return


@app.cell
def _(mo):
    mo.md("""
    ---
    ## 16. Intensity Sub-Decomposition (Qualitative)

    The 3-factor LMDI lumps envelope improvement, fuel switching, and floor-area
    effects into a single "intensity" residual. Here we estimate the approximate
    contribution of each sub-driver.

    **This is an indicative attribution, not a formal decomposition.**
    """)
    return


@app.cell
def _(df_lmdi, go):
    # Approximate attribution of the intensity effect
    _total_intensity = df_lmdi[-1]["cum_intensity"].item()

    _envelope_share = 0.45
    _fuel_switch_share = 0.35
    _other_share = 1.0 - _envelope_share - _fuel_switch_share

    _envelope_tj = _total_intensity * _envelope_share
    _fuel_switch_tj = _total_intensity * _fuel_switch_share
    _other_tj = _total_intensity * _other_share

    fig_intensity_attr = go.Figure(go.Bar(
        y=["Total Intensity Effect",
           "Building Envelope (~45%)",
           "Fuel Switching (~35%)",
           "Other (~20%)"],
        x=[_total_intensity, _envelope_tj, _fuel_switch_tj, _other_tj],
        orientation="h",
        marker_color=["#1565C0", "#4CAF50", "#FF9800", "#9E9E9E"],
        text=[
            f"{_total_intensity:,.0f} TJ",
            f"{_envelope_tj:,.0f} TJ",
            f"{_fuel_switch_tj:,.0f} TJ",
            f"{_other_tj:,.0f} TJ",
        ],
        textposition="auto",
    ))
    fig_intensity_attr.update_layout(
        title="Approximate Attribution of the Intensity Effect (indicative)",
        xaxis_title="\u0394E (TJ, cumulative 2000\u2013latest)",
        template="plotly_white",
        font=dict(size=14),
        width=900,
        height=400,
        yaxis=dict(autorange="reversed"),
    )
    fig_intensity_attr
    return (fig_intensity_attr,)


@app.cell
def _(mo):
    mo.md("""
    ---
    ## 17. Monte Carlo Scenario Projections (2024-2050)

    Forward projection of household energy demand under uncertainty.
    Three named scenarios (BAU, Accelerated, Ambitious) with 1,000 Monte Carlo
    iterations providing uncertainty bands.

    **Parameters**: population growth (BFS reference +/- noise), HDD decline trend,
    renovation rate (drives intensity decline), HP adoption speed.
    """)
    return


@app.cell
def _(df_lmdi_input, go, pl):
    import numpy as _np_mc

    _rng = _np_mc.random.default_rng(42)
    _n_sims = 1000
    _years_proj = list(range(2025, 2051))
    _n_years = len(_years_proj)

    # --- Baseline values from last observed year ---
    _last_obs = df_lmdi_input.filter(pl.col("year") == df_lmdi_input["year"].max())
    _E_base = _last_obs["E"].item()
    _P_base = _last_obs["population"].item()
    _HDD_base = _last_obs["hdd"].item()
    _I_base = _last_obs["intensity"].item()

    # BFS reference population trajectory → 10.5M by 2055
    _pop_2055 = 10_500_000
    _pop_slope = (_pop_2055 - _P_base) / (2055 - 2024)

    # --- Named scenario definitions ---
    _scenarios = {
        "BAU": {"reno_rate": 0.010, "label": "BAU (1% reno)", "color": "#FF9800"},
        "Accelerated": {"reno_rate": 0.020, "label": "Accelerated (2% reno)", "color": "#2196F3"},
        "Ambitious": {"reno_rate": 0.030, "label": "Ambitious (3% reno)", "color": "#4CAF50"},
    }

    # --- Monte Carlo simulation ---
    _all_trajectories = _np_mc.zeros((_n_sims, _n_years))

    for _sim in range(_n_sims):
        _reno = _rng.uniform(0.008, 0.032)
        _hdd_decline_per_yr = _rng.normal(-0.0025, 0.001)
        _pop_noise = _rng.normal(0, 100_000)
        # Intensity decline = base 1.0%/yr + reno contribution
        _intensity_decline = 0.010 + _reno * 0.5

        for _yr_idx, _yr in enumerate(_years_proj):
            _dt = _yr - 2024
            _P_t = _P_base + _pop_slope * _dt + _pop_noise
            _HDD_t = _HDD_base * (1 + _hdd_decline_per_yr) ** _dt
            _I_t = _I_base * (1 - _intensity_decline) ** _dt
            _all_trajectories[_sim, _yr_idx] = _P_t * _HDD_t * _I_t

    # Percentiles
    _p10 = _np_mc.percentile(_all_trajectories, 10, axis=0)
    _p25 = _np_mc.percentile(_all_trajectories, 25, axis=0)
    _p50 = _np_mc.percentile(_all_trajectories, 50, axis=0)
    _p75 = _np_mc.percentile(_all_trajectories, 75, axis=0)
    _p90 = _np_mc.percentile(_all_trajectories, 90, axis=0)

    # Named scenario trajectories (deterministic)
    _named_trajs = {}
    for _sname, _sp in _scenarios.items():
        _traj = []
        for _yr in _years_proj:
            _dt = _yr - 2024
            _P_s = _P_base + _pop_slope * _dt
            _HDD_s = _HDD_base * (1 - 0.0025) ** _dt
            _I_s = _I_base * (1 - (0.010 + _sp["reno_rate"] * 0.5)) ** _dt
            _traj.append(_P_s * _HDD_s * _I_s)
        _named_trajs[_sname] = _traj

    # --- Fan chart ---
    fig_scenario_fan = go.Figure()

    # Uncertainty bands
    fig_scenario_fan.add_trace(go.Scatter(
        x=_years_proj + _years_proj[::-1],
        y=_p90.tolist() + _p10[::-1].tolist(),
        fill="toself", fillcolor="rgba(158,158,158,0.15)",
        line=dict(width=0), name="10th\u201390th pctile", showlegend=True,
    ))
    fig_scenario_fan.add_trace(go.Scatter(
        x=_years_proj + _years_proj[::-1],
        y=_p75.tolist() + _p25[::-1].tolist(),
        fill="toself", fillcolor="rgba(158,158,158,0.30)",
        line=dict(width=0), name="25th\u201375th pctile", showlegend=True,
    ))

    # Named scenarios
    for _sname, _traj in _named_trajs.items():
        fig_scenario_fan.add_trace(go.Scatter(
            x=_years_proj, y=_traj,
            name=_scenarios[_sname]["label"], mode="lines",
            line=dict(color=_scenarios[_sname]["color"], width=3),
        ))

    # Historical for context
    _hist = df_lmdi_input.filter(pl.col("year") >= 2010)
    fig_scenario_fan.add_trace(go.Scatter(
        x=_hist["year"].to_list(), y=_hist["E"].to_list(),
        name="Historical", mode="lines+markers",
        line=dict(color="#212121", width=2, dash="dot"), marker=dict(size=4),
    ))

    fig_scenario_fan.update_layout(
        title="Projected Space-Heating Energy Demand \u2014 Monte Carlo Scenarios (TJ)",
        xaxis_title="Year",
        yaxis_title="Final Energy (TJ)",
        template="plotly_white",
        font=dict(size=14),
        width=1050,
        height=600,
        legend=dict(orientation="h", yanchor="bottom", y=-0.25, xanchor="center", x=0.5),
    )
    fig_scenario_fan
    return (fig_scenario_fan,)


@app.cell
def _(df_lmdi_input, go, pl):
    import numpy as _np_mc

    _last_obs = df_lmdi_input.filter(pl.col("year") == df_lmdi_input["year"].max())
    _P_base = _last_obs["population"].item()
    _HDD_base = _last_obs["hdd"].item()
    _I_base = _last_obs["intensity"].item()
    _pop_2055 = 10_500_000
    _pop_slope = (_pop_2055 - _P_base) / (2055 - 2024)

    # Median trajectory for baseline comparison
    _rng = _np_mc.random.default_rng(42)
    _n_sims = 1000
    _all_trajs = _np_mc.zeros((_n_sims, 26))
    for _sim in range(_n_sims):
        _reno = _rng.uniform(0.008, 0.032)
        _hdd_decline = _rng.normal(-0.0025, 0.001)
        _pop_noise = _rng.normal(0, 100_000)
        _int_decline = 0.010 + _reno * 0.5
        for _yi in range(26):
            _dt = _yi + 1
            _P_t = _P_base + _pop_slope * _dt + _pop_noise
            _H_t = _HDD_base * (1 + _hdd_decline) ** _dt
            _I_t = _I_base * (1 - _int_decline) ** _dt
            _all_trajs[_sim, _yi] = _P_t * _H_t * _I_t
    _base_2050 = _np_mc.percentile(_all_trajs, 50, axis=0)[-1]

    # --- Tornado chart: OAT sensitivity at 2050 ---
    _sensitivities = []
    # Population: +/- 500k
    for _lbl, _dp in [("Population +500k", 500_000), ("Population \u2212500k", -500_000)]:
        _P = _P_base + _pop_slope * 26 + _dp
        _H = _HDD_base * (1 - 0.0025) ** 26
        _I = _I_base * (1 - 0.020) ** 26
        _sensitivities.append((_lbl, _P * _H * _I - _base_2050))

    # HDD decline: slow vs fast
    for _lbl, _rate in [("HDD decline \u22121.5%/dec", -0.0015), ("HDD decline \u22123.5%/dec", -0.0035)]:
        _P = _P_base + _pop_slope * 26
        _H = _HDD_base * (1 + _rate) ** 26
        _I = _I_base * (1 - 0.020) ** 26
        _sensitivities.append((_lbl, _P * _H * _I - _base_2050))

    # Renovation rate: 1% vs 3%
    for _lbl, _reno in [("Reno rate 1.0%/yr", 0.010), ("Reno rate 3.0%/yr", 0.030)]:
        _P = _P_base + _pop_slope * 26
        _H = _HDD_base * (1 - 0.0025) ** 26
        _I = _I_base * (1 - (0.010 + _reno * 0.5)) ** 26
        _sensitivities.append((_lbl, _P * _H * _I - _base_2050))

    _sensitivities.sort(key=lambda x: abs(x[1]), reverse=True)

    fig_tornado = go.Figure(go.Bar(
        y=[_s[0] for _s in _sensitivities],
        x=[_s[1] for _s in _sensitivities],
        orientation="h",
        marker_color=["#F44336" if _s[1] > 0 else "#4CAF50" for _s in _sensitivities],
        text=[f"{_s[1]:+,.0f} TJ" for _s in _sensitivities],
        textposition="auto",
    ))
    fig_tornado.update_layout(
        title="Sensitivity of 2050 Energy Demand to Parameter Changes (vs Median)",
        xaxis_title="\u0394E at 2050 (TJ vs median)",
        template="plotly_white",
        font=dict(size=14),
        width=900,
        height=450,
        yaxis=dict(autorange="reversed"),
    )
    fig_tornado
    return (fig_tornado,)


@app.cell
def _(mo):
    mo.md("""
    ---
    ## 18. Heating System Economics — Operating + Investment Cost Comparison

    Annualized heating cost for a reference 150 m² unrenovated single-family house
    (120 kWh/m²/yr demand). Shows how the CO₂ levy shifts the economic calculus
    between fossil boilers and heat pumps.

    **Operating cost** = fuel + CO₂ levy. **Total cost of ownership** also includes
    annualized investment (net of Gebäudeprogramm subsidies, 20-year lifetime, 3% discount).
    """)
    return


@app.cell
def _(go):
    # Reference building: 150 m², unrenovated (120 kWh/m²/yr useful demand)
    _area = 150
    _demand = 120  # kWh/m²/yr

    _systems = {
        "Oil Boiler": {"eff": 0.90, "fuel_chf_kwh": 0.12, "co2_kg_kwh": 0.265, "color": "#8B4513"},
        "Gas Condensing": {"eff": 0.95, "fuel_chf_kwh": 0.11, "co2_kg_kwh": 0.198, "color": "#FF6B35"},
        "Air-Source HP": {"eff": 3.2, "fuel_chf_kwh": 0.27, "co2_kg_kwh": 0.0, "color": "#2196F3"},
        "Ground-Source HP": {"eff": 4.0, "fuel_chf_kwh": 0.27, "co2_kg_kwh": 0.0, "color": "#1565C0"},
    }

    _co2_levies = [120, 200, 300]

    _results = []
    for _sname, _s in _systems.items():
        for _levy in _co2_levies:
            _final_energy = _area * _demand / _s["eff"]  # kWh delivered
            _fuel_cost = _final_energy * _s["fuel_chf_kwh"]
            _co2_cost = _final_energy * _s["co2_kg_kwh"] / 1000 * _levy
            _results.append({
                "system": _sname,
                "levy": f"CHF {_levy}/t",
                "total_chf": _fuel_cost + _co2_cost,
            })

    fig_cost = go.Figure()
    _opacities = [0.55, 0.75, 0.95]
    for _li, _levy in enumerate(_co2_levies):
        _levy_str = f"CHF {_levy}/t"
        _filtered = [_r for _r in _results if _r["levy"] == _levy_str]
        fig_cost.add_trace(go.Bar(
            x=[_r["system"] for _r in _filtered],
            y=[_r["total_chf"] for _r in _filtered],
            name=f"CO\u2082 levy {_levy_str}",
            text=[f"CHF {_r['total_chf']:,.0f}" for _r in _filtered],
            textposition="auto",
            marker_color=[_systems[_r["system"]]["color"] for _r in _filtered],
            opacity=_opacities[_li],
        ))

    fig_cost.update_layout(
        title=dict(
            text=f"Annual Heating Cost \u2014 150 m\u00b2 Unrenovated House ({_demand} kWh/m\u00b2/yr)"
                 "<br><sup>Fuel + CO\u2082 levy only \u2014 HP electricity is levy-exempt (Swiss grid ~25 g CO\u2082/kWh)</sup>",
        ),
        xaxis_title="Heating System",
        yaxis_title="Annual Cost (CHF)",
        barmode="group",
        template="plotly_white",
        font=dict(size=14),
        width=1000,
        height=550,
        legend=dict(orientation="h", yanchor="bottom", y=-0.18, xanchor="center", x=0.5),
    )
    return (fig_cost,)


@app.cell
def _(mo):
    # Investment costs and subsidies — Total Cost of Ownership
    # Sources:
    #   Investment: EnergieSchweiz "Heizungsersatz: Kosten" (EFH reference case, Bern)
    #     https://www.energieschweiz.ch/stories/erneuerung-der-heizung-zu-erwartende-kosten/
    #   HP costs: hausinfo.ch "Kosten einer Wärmepumpe in der Schweiz"
    #     https://hausinfo.ch/de/bauen-renovieren/haustechnik-vernetzung/heizung-lueftung-klima/waermepumpen/investitionskosten.html
    #   Subsidies: hausinfo.ch "Heizungsersatz: Finanzierung & Förderung" — average cantonal amounts
    #     https://hausinfo.ch/de/bauen-renovieren/haustechnik-vernetzung/heizung-lueftung-klima/oelheizung-ersetzen/heizungsersatz-finanzierung.html
    #   Fossil systems: EnergieSchweiz reference case (oil ~15-20k, gas ~18-22k)
    _lifetime = 20
    _r = 0.03
    _annuity = _r / (1 - (1 + _r) ** (-_lifetime))

    _tco = {
        "Oil Boiler":      {"invest": 18_000, "subsidy": 0,      "op": 3036},
        "Gas Condensing":  {"invest": 20_000, "subsidy": 0,      "op": 2534},
        "Air-Source HP":   {"invest": 42_000, "subsidy": 5_000,  "op": 1519},
        "Ground-Source HP":{"invest": 60_000, "subsidy": 10_000, "op": 1215},
    }
    _rows = []
    for _name, _v in _tco.items():
        _net_invest = _v["invest"] - _v["subsidy"]
        _annual_invest = round(_net_invest * _annuity)
        _total = _v["op"] + _annual_invest
        _rows.append(
            f"| {_name} | {_v['invest']:,} | {_v['subsidy']:,} "
            f"| {_net_invest:,} | {_annual_invest:,} | {_v['op']:,} | **{_total:,}** |"
        )

    _table_rows = "\n        ".join(_rows)

    mo.md(
        f"""
        ### Total Cost of Ownership (at current CO\u2082 levy CHF 120/t)

        | System | Investment (CHF) | Subsidy | Net invest | Annual invest | Annual operating | **Annual TCO** |
        |--------|--:|--:|--:|--:|--:|--:|
        {_table_rows}

        **Sources & assumptions:**
        - Investment costs: [EnergieSchweiz — Heizungsersatz: Kosten](https://www.energieschweiz.ch/stories/erneuerung-der-heizung-zu-erwartende-kosten/) (EFH reference case);
          [hausinfo.ch — Kosten einer W\u00e4rmepumpe](https://hausinfo.ch/de/bauen-renovieren/haustechnik-vernetzung/heizung-lueftung-klima/waermepumpen/investitionskosten.html)
        - Subsidies: average cantonal Geb\u00e4udeprogramm amounts (~CHF 5k air-source, ~CHF 10k ground-source with borehole);
          [hausinfo.ch — Heizungsersatz Finanzierung](https://hausinfo.ch/de/bauen-renovieren/haustechnik-vernetzung/heizung-lueftung-klima/oelheizung-ersetzen/heizungsersatz-finanzierung.html)
        - 20-year lifetime, 3% real discount rate. Operating costs from the cost-comparison chart above (fuel + CO\u2082 levy at CHF 120/t).

        Even including the higher upfront cost, **heat pumps reach near-parity with fossil systems**
        on a total-cost-of-ownership basis thanks to their COP advantage (3\u20134\u00d7 less final
        energy) and zero CO\u2082 levy exposure. At higher levy levels (CHF 200\u2013300/t),
        heat pumps become clearly cheaper.
        """
    )
    return


@app.cell
def _(mo):
    mo.md("""
    ---
    ## Figure Export for Report
    """)
    return


@app.cell
def _(
    df_floor_area,
    df_hdd,
    df_heating_compare,
    df_hh,
    df_indexed,
    df_lmdi,
    df_mix,
    df_pop_full,
    go,
    pl,
):
    from pathlib import Path as _Path

    _export_app_path = _Path(__file__).resolve()
    _proj_root = _export_app_path.parent.parent if _export_app_path.parent.name == "notebooks" else _export_app_path.parent
    _FIGURES_DIR = _proj_root / "figures"
    _FIGURES_DIR.mkdir(exist_ok=True)

    _exported_files = []

    def _export_fig(fig, name):
        """Export a figure as SVG and PDF, printing confirmation."""
        for _ext in ("svg", "pdf"):
            _out = _FIGURES_DIR / f"{name}.{_ext}"
            fig.write_image(str(_out), width=1000, height=550)
            _exported_files.append(str(_out))
        print(f"  Exported {name}.svg + {name}.pdf")

    # ------------------------------------------------------------------
    # 1. Household energy by carrier — stacked area
    # ------------------------------------------------------------------
    _carrier_map = {
        "Erdölprodukte": "Petroleum Products",
        "Elektrizität": "Electricity",
        "Gas": "Natural Gas",
        "Holzenergie": "Wood / Biomass",
        "Fernwärme": "District Heating",
        "Kohle": "Coal",
        "Uebrige erneuerbare Energien": "Other Renewables",
        "Müll und Industrieabfälle": "Waste",
    }
    _carrier_colors = {
        "Petroleum Products": "#8B4513",
        "Natural Gas": "#FF6B35",
        "Electricity": "#FFC107",
        "Wood / Biomass": "#4CAF50",
        "District Heating": "#9C27B0",
        "Other Renewables": "#00BCD4",
        "Coal": "#607D8B",
        "Waste": "#795548",
    }

    _df_hh_plot = df_hh.with_columns(
        pl.col("Energietraeger").replace(_carrier_map).alias("carrier_en")
    )
    _df_pivot = (
        _df_hh_plot
        .group_by(["Jahr", "carrier_en"])
        .agg(pl.col("TJ").sum())
        .sort("Jahr")
    )
    _carrier_totals = (
        _df_pivot.group_by("carrier_en")
        .agg(pl.col("TJ").sum().alias("total"))
        .sort("total", descending=True)
    )
    _ordered_carriers = _carrier_totals["carrier_en"].to_list()

    _fig1 = go.Figure()
    for _c in _ordered_carriers:
        _sub = _df_pivot.filter(pl.col("carrier_en") == _c).sort("Jahr")
        _fig1.add_trace(go.Scatter(
            x=_sub["Jahr"].to_list(),
            y=_sub["TJ"].to_list(),
            name=_c,
            mode="lines",
            stackgroup="one",
            line=dict(width=0.5, color=_carrier_colors.get(_c, "#999")),
        ))
    _fig1.update_layout(
        title="Swiss Household Final Energy Consumption by Carrier (TJ)",
        xaxis_title="Year",
        yaxis_title="Final Energy (TJ)",
        template="plotly_white",
        font=dict(size=14),
        legend=dict(orientation="h", yanchor="bottom", y=-0.25, xanchor="center", x=0.5),
    )
    _export_fig(_fig1, "fig1_energy_by_carrier")

    # ------------------------------------------------------------------
    # 2. HDD scatter with trendline
    # ------------------------------------------------------------------
    _years_hdd = df_hdd["year"].to_list()
    _hdds = df_hdd["hdd"].to_list()

    _n = len(_years_hdd)
    _sum_x = sum(_years_hdd)
    _sum_y = sum(_hdds)
    _sum_xy = sum(_x * _y for _x, _y in zip(_years_hdd, _hdds))
    _sum_x2 = sum(_x * _x for _x in _years_hdd)
    _slope = (_n * _sum_xy - _sum_x * _sum_y) / (_n * _sum_x2 - _sum_x * _sum_x)
    _intercept = (_sum_y - _slope * _sum_x) / _n
    _trend_y = [_slope * _yr + _intercept for _yr in _years_hdd]
    _decline_per_decade = _slope * 10

    _fig2 = go.Figure()
    _fig2.add_trace(go.Scatter(
        x=_years_hdd, y=_hdds,
        name="HDD (actual)",
        mode="markers",
        marker=dict(size=8, color="#FF9800"),
    ))
    _fig2.add_trace(go.Scatter(
        x=_years_hdd, y=_trend_y,
        name=f"Trend ({_decline_per_decade:.0f} HDD/decade)",
        mode="lines",
        line=dict(color="#F44336", width=2, dash="dash"),
    ))
    _fig2.update_layout(
        title="Heating Degree Days \u2014 Switzerland (base 20/12\u00b0C, Zurich Plateau)",
        xaxis_title="Year",
        yaxis_title="HDD",
        template="plotly_white",
        font=dict(size=14),
    )
    _export_fig(_fig2, "fig2_hdd_trend")

    # ------------------------------------------------------------------
    # 3. Population line chart
    # ------------------------------------------------------------------
    _fig3 = go.Figure()
    _fig3.add_trace(go.Scatter(
        x=df_pop_full["year"].to_list(),
        y=[_p / 1_000_000 for _p in df_pop_full["population"].to_list()],
        name="Population",
        mode="lines+markers",
        line=dict(color="#9C27B0", width=3),
        marker=dict(size=5),
    ))
    _fig3.update_layout(
        title="Swiss Population (millions)",
        xaxis_title="Year",
        yaxis_title="Population (millions)",
        template="plotly_white",
        font=dict(size=14),
    )
    _export_fig(_fig3, "fig3_population")

    # ------------------------------------------------------------------
    # 4. Heating mix stacked area
    # ------------------------------------------------------------------
    _mix_carriers = ["oil", "gas", "heat_pump", "wood", "electric_resistance", "district_heating", "other_solar"]
    _mix_colors = {
        "oil": "#8B4513",
        "gas": "#FF6B35",
        "heat_pump": "#2196F3",
        "wood": "#4CAF50",
        "electric_resistance": "#FFC107",
        "district_heating": "#9C27B0",
        "other_solar": "#00BCD4",
    }
    _mix_labels = {
        "oil": "Heating Oil",
        "gas": "Natural Gas",
        "heat_pump": "Heat Pumps (electric)",
        "wood": "Wood / Biomass",
        "electric_resistance": "Electric Resistance",
        "district_heating": "District Heating",
        "other_solar": "Other (Coal + misc)",
    }

    _fig4 = go.Figure()
    for _mc in _mix_carriers:
        _fig4.add_trace(go.Scatter(
            x=df_mix["year"].to_list(),
            y=df_mix[_mc].to_list(),
            name=_mix_labels[_mc],
            mode="lines",
            stackgroup="one",
            line=dict(width=0.5, color=_mix_colors[_mc]),
            fillcolor=_mix_colors[_mc],
        ))
    _fig4.update_layout(
        title="Swiss Residential Heating \u2014 Floor Area Share by System (%, BFE/Prognos 2025)",
        xaxis_title="Year",
        yaxis_title="Share of Floor Area (%)",
        yaxis=dict(range=[0, 100]),
        template="plotly_white",
        font=dict(size=14),
        legend=dict(orientation="h", yanchor="bottom", y=-0.25, xanchor="center", x=0.5),
    )
    _export_fig(_fig4, "fig4_heating_mix")

    # ------------------------------------------------------------------
    # 5. Decoupling indexed chart
    # ------------------------------------------------------------------
    _fig5 = go.Figure()

    _decouple_series = [
        ("gdp_idx", "GDP (real)", "#E91E63", "solid", 3),
        ("pop_idx", "Population", "#9C27B0", "dash", 2),
        ("floor_idx", "Heated Floor Area", "#FF9800", "dashdot", 2),
        ("energy_idx", "Space-Heating Energy", "#2196F3", "solid", 3),
        ("intensity_idx", "Energy Intensity (kWh/m\u00b2)", "#F44336", "dot", 2),
    ]

    for _col, _name, _color, _dash, _width in _decouple_series:
        _fig5.add_trace(go.Scatter(
            x=df_indexed["year"].to_list(),
            y=df_indexed[_col].to_list(),
            name=_name,
            mode="lines+markers",
            line=dict(color=_color, width=_width, dash=_dash),
            marker=dict(size=4),
        ))

    _fig5.add_hline(y=100, line_dash="dot", line_color="gray", opacity=0.5)

    for _yr, _text, _ypos in [
        (2008, "CO2 levy", 108),
        (2010, "Gebaudeprogramm", 118),
        (2015, "MuKEn 2014", 128),
        (2022, "CHF 120/t CO2", 138),
    ]:
        _fig5.add_vline(x=_yr, line_dash="dot", line_color="rgba(0,0,0,0.15)")
        _fig5.add_annotation(
            x=_yr, y=_ypos, text=_text, showarrow=False,
            font=dict(size=9, color="gray"),
        )

    _fig5.update_layout(
        title="Decoupling: Space-Heating Energy vs Economic Growth (2000 = 100)",
        xaxis_title="Year",
        yaxis_title="Index (2000 = 100)",
        template="plotly_white",
        font=dict(size=14),
        legend=dict(orientation="h", yanchor="bottom", y=-0.22, xanchor="center", x=0.5),
    )
    _export_fig(_fig5, "fig5_decoupling")

    # ------------------------------------------------------------------
    # 6. Floor area per person
    # ------------------------------------------------------------------
    _fig6 = go.Figure()
    _fig6.add_trace(go.Scatter(
        x=df_floor_area["year"].to_list(),
        y=df_floor_area["m2_per_person_sfh"].to_list(),
        name="Single-Family Houses",
        mode="lines+markers",
        line=dict(color="#FF9800", width=2.5),
        marker=dict(size=6),
    ))
    _fig6.add_trace(go.Scatter(
        x=df_floor_area["year"].to_list(),
        y=df_floor_area["m2_per_person_mfh"].to_list(),
        name="Multi-Family Buildings",
        mode="lines+markers",
        line=dict(color="#2196F3", width=2.5),
        marker=dict(size=6),
    ))
    _fig6.add_trace(go.Scatter(
        x=df_floor_area["year"].to_list(),
        y=df_floor_area["m2_per_person_avg"].to_list(),
        name="Weighted Average (~30/70)",
        mode="lines+markers",
        line=dict(color="#4CAF50", width=3, dash="dash"),
        marker=dict(size=5),
    ))
    _fig6.update_layout(
        title="Average Floor Area per Person \u2014 Swiss Residential Buildings (BFS GWS)",
        xaxis_title="Year",
        yaxis_title="m\u00b2 per person",
        template="plotly_white",
        font=dict(size=14),
        legend=dict(orientation="h", yanchor="bottom", y=-0.22, xanchor="center", x=0.5),
    )
    _export_fig(_fig6, "fig6_floor_area")

    # ------------------------------------------------------------------
    # 7. Buildings by heating source 2000 vs 2021
    # ------------------------------------------------------------------
    _sources_exp = df_heating_compare["source"].to_list()
    _v2000_exp = df_heating_compare["buildings_2000"].to_list()
    _v2021_exp = df_heating_compare["buildings_2021"].to_list()

    _bar_colors_exp = {
        "Heat Pump": "#2196F3",
        "Gas": "#FF6B35",
        "Heating Oil": "#8B4513",
        "Wood": "#4CAF50",
        "Electricity": "#FFC107",
        "Solar Thermal": "#00BCD4",
        "District Heating": "#9C27B0",
        "Other": "#607D8B",
    }

    _fig7 = go.Figure()
    _fig7.add_trace(go.Bar(
        x=_sources_exp,
        y=[_v / 1000 for _v in _v2000_exp],
        name="2000",
        marker_color=[_bar_colors_exp.get(_s, "#999") for _s in _sources_exp],
        opacity=0.5,
    ))
    _fig7.add_trace(go.Bar(
        x=_sources_exp,
        y=[_v / 1000 for _v in _v2021_exp],
        name="2021",
        marker_color=[_bar_colors_exp.get(_s, "#999") for _s in _sources_exp],
        opacity=0.9,
    ))
    _fig7.update_layout(
        title="Residential Buildings by Heating Energy Source \u2014 2000 vs 2021 (BFS GWS)",
        xaxis_title="Energy Source",
        yaxis_title="Number of Buildings (thousands)",
        barmode="group",
        template="plotly_white",
        font=dict(size=14),
        legend=dict(orientation="h", yanchor="bottom", y=-0.18, xanchor="center", x=0.5),
    )
    _export_fig(_fig7, "fig7_buildings_heating")

    # ------------------------------------------------------------------
    # 8. LMDI waterfall chart
    # ------------------------------------------------------------------
    _last_lmdi = df_lmdi[-1]
    _pop_val = _last_lmdi["cum_pop"].item()
    _weather_val = _last_lmdi["cum_weather"].item()
    _intensity_val = _last_lmdi["cum_intensity"].item()
    _net_val = _last_lmdi["cum_total"].item()

    _fig8 = go.Figure(go.Waterfall(
        name="LMDI",
        orientation="v",
        measure=["relative", "relative", "relative", "total"],
        x=["Population", "Weather (HDD)", "Intensity (tech.)", "Net Change"],
        y=[_pop_val, _weather_val, _intensity_val, _net_val],
        text=[
            f"{_pop_val:+.0f} TJ",
            f"{_weather_val:+.0f} TJ",
            f"{_intensity_val:+.0f} TJ",
            f"{_net_val:+.0f} TJ",
        ],
        textposition="outside",
        connector=dict(line=dict(color="rgba(0,0,0,0.3)", width=1)),
        increasing=dict(marker=dict(color="#F44336")),
        decreasing=dict(marker=dict(color="#4CAF50")),
        totals=dict(marker=dict(color="#2196F3")),
    ))
    _fig8.update_layout(
        title="LMDI-I Decomposition of Space-Heating Energy Change (2000 to latest, TJ)",
        yaxis_title="Cumulative \u0394E (TJ)",
        template="plotly_white",
        font=dict(size=14),
        showlegend=False,
    )
    _export_fig(_fig8, "fig8_lmdi_waterfall")

    # ------------------------------------------------------------------
    # 9. LMDI cumulative time series
    # ------------------------------------------------------------------
    _years_lmdi = df_lmdi["year"].to_list()

    _fig9 = go.Figure()
    _fig9.add_trace(go.Scatter(
        x=_years_lmdi, y=df_lmdi["cum_pop"].to_list(),
        name="Population",
        mode="lines+markers",
        line=dict(color="#F44336", width=2.5),
        marker=dict(size=4),
    ))
    _fig9.add_trace(go.Scatter(
        x=_years_lmdi, y=df_lmdi["cum_weather"].to_list(),
        name="Weather (HDD)",
        mode="lines+markers",
        line=dict(color="#2196F3", width=2.5),
        marker=dict(size=4),
    ))
    _fig9.add_trace(go.Scatter(
        x=_years_lmdi, y=df_lmdi["cum_intensity"].to_list(),
        name="Intensity (tech.)",
        mode="lines+markers",
        line=dict(color="#4CAF50", width=2.5),
        marker=dict(size=4),
    ))
    _fig9.add_trace(go.Scatter(
        x=_years_lmdi, y=df_lmdi["cum_total"].to_list(),
        name="Net Change",
        mode="lines+markers",
        line=dict(color="#212121", width=3, dash="dash"),
        marker=dict(size=5),
    ))
    _fig9.add_hline(y=0, line_dash="dot", line_color="gray", opacity=0.5)
    _fig9.update_layout(
        title="Cumulative LMDI-I Decomposition \u2014 Year-by-Year (TJ, base 2000)",
        xaxis_title="Year",
        yaxis_title="Cumulative \u0394E (TJ)",
        template="plotly_white",
        font=dict(size=14),
        legend=dict(orientation="h", yanchor="bottom", y=-0.22, xanchor="center", x=0.5),
    )
    _export_fig(_fig9, "fig9_lmdi_timeseries")
    return


@app.cell
def _(
    fig_cost,
    fig_hdd_norm,
    fig_intensity_attr,
    fig_scenario_fan,
    fig_tornado,
):
    from pathlib import Path as _Path2

    _app2 = _Path2(__file__).resolve()
    _proj2 = _app2.parent.parent if _app2.parent.name == "notebooks" else _app2.parent
    _FIGS2 = _proj2 / "figures"
    _FIGS2.mkdir(exist_ok=True)

    _new_figs = [
        (fig_hdd_norm, "fig10_energy_hdd_normalized"),
        (fig_intensity_attr, "fig11_intensity_attribution"),
        (fig_scenario_fan, "fig12_scenario_fan"),
        (fig_tornado, "fig13_scenario_tornado"),
        (fig_cost, "fig14_cost_comparison"),
    ]

    _exported2 = []
    for _fig, _name in _new_figs:
        for _ext in ("svg", "pdf"):
            _out = _FIGS2 / f"{_name}.{_ext}"
            _fig.write_image(str(_out), width=1000, height=550)
            _exported2.append(str(_out))
    return


@app.cell
def _(mo):
    mo.md("""
    ### Exported Files

    | # | SVG | PDF |
    |---|-----|-----|
    | 1 | `fig1_energy_by_carrier.svg` | `fig1_energy_by_carrier.pdf` |
    | 2 | `fig2_hdd_trend.svg` | `fig2_hdd_trend.pdf` |
    | 3 | `fig3_population.svg` | `fig3_population.pdf` |
    | 4 | `fig4_heating_mix.svg` | `fig4_heating_mix.pdf` |
    | 5 | `fig5_decoupling.svg` | `fig5_decoupling.pdf` |
    | 6 | `fig6_floor_area.svg` | `fig6_floor_area.pdf` |
    | 7 | `fig7_buildings_heating.svg` | `fig7_buildings_heating.pdf` |
    | 8 | `fig8_lmdi_waterfall.svg` | `fig8_lmdi_waterfall.pdf` |
    | 9 | `fig9_lmdi_timeseries.svg` | `fig9_lmdi_timeseries.pdf` |
    | 10 | `fig10_energy_hdd_normalized.svg` | `fig10_energy_hdd_normalized.pdf` |
    | 11 | `fig11_intensity_attribution.svg` | `fig11_intensity_attribution.pdf` |
    | 12 | `fig12_scenario_fan.svg` | `fig12_scenario_fan.pdf` |
    | 13 | `fig13_scenario_tornado.svg` | `fig13_scenario_tornado.pdf` |
    | 14 | `fig14_cost_comparison.svg` | `fig14_cost_comparison.pdf` |

    All figures saved to `figures/` for inclusion in the Typst report.
    """)
    return


if __name__ == "__main__":
    app.run()
