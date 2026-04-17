"""Data download, caching, and parsing functions.

All raw data used by the HUM-470 notebook is fetched through helpers in this
module so that downloads are cached and parsing logic is reusable.
"""

from __future__ import annotations

import json
from pathlib import Path

import httpx
import polars as pl


# ────────────────────────────────────────────────────────────────────────
# Generic download helpers
# ────────────────────────────────────────────────────────────────────────

def download_file(
    directory: Path,
    url: str,
    filename: str,
    *,
    force: bool = False,
) -> Path:
    """Download *url* into *directory/filename* if not already cached."""
    path = directory / filename
    if path.exists() and not force:
        print(f"  [cached] {filename}")
        return path
    print(f"  [downloading] {filename} ...")
    resp = httpx.get(url, follow_redirects=True, timeout=60.0)
    resp.raise_for_status()
    path.write_bytes(resp.content)
    print(f"  [saved] {filename} ({len(resp.content) / 1024:.0f} KB)")
    return path


def download_json(
    directory: Path,
    url: str,
    filename: str,
    *,
    payload: dict | None = None,
    force: bool = False,
) -> Path:
    """Download JSON from a REST API, cache locally in *directory*."""
    path = directory / filename
    if path.exists() and not force:
        print(f"  [cached] {filename}")
        return path
    print(f"  [downloading] {filename} ...")
    if payload:
        resp = httpx.post(url, json=payload, follow_redirects=True, timeout=30.0)
    else:
        resp = httpx.get(url, follow_redirects=True, timeout=30.0)
    resp.raise_for_status()
    path.write_text(json.dumps(resp.json(), indent=2, ensure_ascii=False))
    print(f"  [saved] {filename} ({len(resp.content) / 1024:.0f} KB)")
    return path


# ────────────────────────────────────────────────────────────────────────
# BFE Energy Balance
# ────────────────────────────────────────────────────────────────────────

def load_household_energy(energy_balance_path: Path) -> pl.DataFrame:
    """Load BFE energy balance and filter for household final consumption.

    Returns a DataFrame with columns: ``Jahr``, ``Energietraeger``, ``TJ``.
    """
    df = (
        pl.read_csv(energy_balance_path, separator=",", infer_schema_length=10_000)
        .with_columns(pl.col("TJ").cast(pl.Float64, strict=False))
        .filter(pl.col("Rubrik") == "Endverbrauch - Haushalte")
        .filter(pl.col("TJ").is_not_null())
        .sort("Jahr", "Energietraeger")
    )
    return df


# ────────────────────────────────────────────────────────────────────────
# BFE/Prognos Space‐Heating Share (Tabelle 1)
# ────────────────────────────────────────────────────────────────────────

def load_sh_share(xlsx_path: Path) -> pl.DataFrame:
    """Parse space‐heating share from BFE/Prognos Tabelle 1.

    Returns DataFrame with columns: ``year``, ``raumwaerme_pj``,
    ``total_hh_pj``, ``sh_share``.
    """
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["Tabelle1"]

    records: list[dict] = []
    for c in range(3, 28):  # col 3 = year 2000, col 27 = year 2024
        yr = int(ws.cell(6, c).value)
        rw_pj = float(ws.cell(7, c).value)   # Raumwärme (PJ)
        tot_pj = float(ws.cell(17, c).value)  # Total HH (PJ)
        records.append({
            "year": yr,
            "raumwaerme_pj": round(rw_pj, 2),
            "total_hh_pj": round(tot_pj, 2),
            "sh_share": round(rw_pj / tot_pj, 4),
        })

    wb.close()
    return pl.DataFrame(records).sort("year")


def get_sh_share_fn(df_sh_share: pl.DataFrame):
    """Return a closure that looks up the SH share for a given year."""
    def _get(year: int) -> float:
        row = df_sh_share.filter(pl.col("year") == year)
        if row.shape[0] > 0:
            return row["sh_share"][0]
        return df_sh_share.filter(
            pl.col("year") == df_sh_share["year"].max()
        )["sh_share"][0]
    return _get


# ────────────────────────────────────────────────────────────────────────
# BFE/Prognos Floor Area by Heating System (Tabelle 13)
# ────────────────────────────────────────────────────────────────────────

_TABELLE13_LABEL_MAP = {
    "Heizöl": "oil",
    "Erdgas": "gas",
    "Elektrische Wärmepumpen": "heat_pump",
    "Holz": "wood",
    "El. Widerstandsheizungen": "electric_resistance",
    "Fernwärme": "district_heating",
    "Kohle": "coal",
    "Übrige/ ohne Heizsystem": "other",
}


def load_heating_mix(xlsx_path: Path) -> pl.DataFrame:
    """Parse floor‐area shares by heating system from Tabelle 13.

    Returns DataFrame with ``year`` and percentage columns for each system.
    Coal + Other are merged into ``other_solar``.
    """
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["Tabelle13"]

    header = list(ws.iter_rows(min_row=6, max_row=6, values_only=True))[0]
    years = [int(v) for v in header[2:27] if isinstance(v, (int, float))]

    system_rows: dict[str, list[float]] = {}
    for row in ws.iter_rows(min_row=7, max_row=14, values_only=True):
        label = row[1]
        if label and label != "Total":
            vals = [float(v) if v is not None else 0.0 for v in row[2:27]]
            system_rows[label.strip()] = vals

    data: dict[str, list] = {"year": years}
    for de_label, en_col in _TABELLE13_LABEL_MAP.items():
        data[en_col] = system_rows.get(de_label, [0.0] * len(years))

    df = pl.DataFrame(data)

    carrier_cols = list(_TABELLE13_LABEL_MAP.values())
    df = df.with_columns(
        _total=sum(pl.col(c) for c in carrier_cols),
    ).with_columns(
        **{c: (pl.col(c) / pl.col("_total") * 100).round(1) for c in carrier_cols}
    ).drop("_total")

    df = df.with_columns(
        other_solar=(pl.col("coal") + pl.col("other")).round(1),
    ).drop("coal", "other")

    wb.close()
    return df


# ────────────────────────────────────────────────────────────────────────
# BFE/Prognos Space‐Heating Energy by Carrier (Tabelle 11)
# ────────────────────────────────────────────────────────────────────────

_TABELLE11_LABEL_MAP = {
    "Heizöl": "oil",
    "Erdgas": "gas",
    "Elektrische Wärmepumpen": "heat_pump_elec",
    "Holz": "wood",
    "El. Widerstandsheizungen": "electric_resistance",
    "Fernwärme": "district_heating",
    "Kohle": "coal",
    "Umweltwärme": "ambient_heat",
    "Solar": "solar",
}


def load_heating_energy(xlsx_path: Path) -> pl.DataFrame:
    """Parse space‐heating energy (PJ) by carrier from Tabelle 11.

    Returns DataFrame with ``year`` and percentage columns for each carrier.
    """
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["Tabelle11"]

    header = list(ws.iter_rows(min_row=6, max_row=6, values_only=True))[0]
    years = [int(v) for v in header[2:27] if isinstance(v, (int, float))]

    carrier_rows: dict[str, list[float]] = {}
    for row in ws.iter_rows(min_row=7, max_row=16, values_only=True):
        label = row[1]
        if label and label != "Total" and not str(label).strip().startswith("darunter"):
            vals = [float(v) if v is not None else 0.0 for v in row[2:27]]
            carrier_rows[label.strip()] = vals

    data: dict[str, list] = {"year": years}
    for de_label, en_col in _TABELLE11_LABEL_MAP.items():
        data[en_col] = carrier_rows.get(de_label, [0.0] * len(years))

    df = pl.DataFrame(data)
    energy_cols = list(_TABELLE11_LABEL_MAP.values())
    df = df.with_columns(
        _total=sum(pl.col(c) for c in energy_cols),
    ).with_columns(
        **{c: (pl.col(c) / pl.col("_total") * 100).round(1) for c in energy_cols}
    ).drop("_total")

    wb.close()
    return df


# ────────────────────────────────────────────────────────────────────────
# HDD — Composite National Series
# ────────────────────────────────────────────────────────────────────────

def load_hdd(hdd_bfe_path: Path, hdd_zurich_path: Path) -> pl.DataFrame:
    """Build composite HDD series: Zurich Fluntern (pre-1994) + BFE national (1994+).

    Returns DataFrame with columns: ``year``, ``hdd``.
    """
    # BFE national (1994+)
    df_bfe = (
        pl.read_csv(hdd_bfe_path, separator=",", infer_schema_length=1000)
        .filter(pl.col("Jahr").is_not_null())
        .with_columns(
            pl.col("Jahr").cast(pl.Int64),
            pl.col("Heizgradtage").cast(pl.Float64, strict=False),
        )
        .filter(pl.col("Heizgradtage").is_not_null())
        .group_by("Jahr")
        .agg(pl.col("Heizgradtage").sum().round(1).alias("hdd"))
        .rename({"Jahr": "year"})
        .sort("year")
    )

    # Zurich Fluntern (pre-1994)
    df_zurich = (
        pl.read_csv(
            hdd_zurich_path,
            separator=",",
            encoding="utf8-lossy",
            infer_schema_length=5000,
        )
        .filter(pl.col("Standort") == "Fluntern")
        .filter(pl.col("Heizgradtag").is_not_null())
        .with_columns(
            pl.col("Jahr").cast(pl.Int64),
            pl.col("Heizgradtag").cast(pl.Float64, strict=False),
        )
        .filter(pl.col("Heizgradtag").is_not_null())
        .group_by("Jahr")
        .agg(pl.col("Heizgradtag").sum().round(1).alias("hdd"))
        .rename({"Jahr": "year"})
        .sort("year")
    )

    df_pre1994 = (
        df_zurich
        .filter(pl.col("year") < 1994)
        .filter(pl.col("year") >= 1980)
    )

    return pl.concat([df_pre1994, df_bfe]).sort("year").unique("year")


# ────────────────────────────────────────────────────────────────────────
# Population — Composite BFS + World Bank (1960–2024)
# ────────────────────────────────────────────────────────────────────────

def load_population(
    bfs_pop_path: Path,
    wb_pop_path: Path,
) -> pl.DataFrame:
    """Build composite population series: World Bank (1960–2009) + BFS STATPOP (2010+).

    BFS STATPOP is the authoritative source for 2010+.  World Bank fills
    the earlier years (1960–2009) and enables the long-run view that shows
    Switzerland's immigration-driven population acceleration.

    Returns DataFrame with columns: ``year``, ``population``.
    """
    # BFS STATPOP (2010+)
    bfs_data = json.loads(bfs_pop_path.read_text())
    bfs_records = [
        {"year": int(e["key"][0]), "population": int(e["values"][0])}
        for e in bfs_data.get("data", [])
        if e["values"][0] and e["values"][0] != "..."
    ]
    df_bfs = pl.DataFrame(bfs_records).sort("year")
    bfs_min_year: int = int(df_bfs["year"].min())  # type: ignore[arg-type]

    # World Bank (fills years before BFS coverage)
    wb_raw = json.loads(wb_pop_path.read_text())
    wb_records = [
        {"year": int(e["date"]), "population": int(e["value"])}
        for e in wb_raw[1]
        if e["value"] is not None and int(e["date"]) < bfs_min_year
    ]
    df_wb = pl.DataFrame(wb_records)

    return pl.concat([df_wb, df_bfs]).sort("year").unique("year")


def load_population_projections(csv_path: Path) -> pl.DataFrame | None:
    """Load BFS population projection scenarios (reference, high, low).

    Returns DataFrame with columns: ``year``, ``pop_reference``,
    ``pop_high``, ``pop_low``.  Returns None if the file does not exist.
    """
    if not csv_path.exists():
        return None
    return pl.read_csv(csv_path, separator=",", infer_schema_length=1000)


# ────────────────────────────────────────────────────────────────────────
# World Bank GDP
# ────────────────────────────────────────────────────────────────────────

def load_gdp(wb_gdp_path: Path) -> pl.DataFrame:
    """Parse World Bank GDP (constant LCU) for Switzerland.

    Returns DataFrame with columns: ``year``, ``gdp_real_chf``, ``gdp_bn_chf``.
    """
    raw = json.loads(wb_gdp_path.read_text())
    records = [
        {"year": int(e["date"]), "gdp_real_chf": float(e["value"])}
        for e in raw[1]
        if e["value"] is not None and 2000 <= int(e["date"]) <= 2024
    ]
    df = pl.DataFrame(records).sort("year")
    return df.with_columns(
        (pl.col("gdp_real_chf") / 1e9).round(1).alias("gdp_bn_chf"),
    )


# ────────────────────────────────────────────────────────────────────────
# BFS Floor Area per Person
# ────────────────────────────────────────────────────────────────────────

def load_floor_area(xlsx_path: Path) -> pl.DataFrame:
    """Parse BFS floor area per person by building category (2012+).

    Returns DataFrame with columns: ``year``, ``m2_per_person_sfh``,
    ``m2_per_person_mfh``, ``m2_per_person_avg``.
    """
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    records: list[dict] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        yr = int(sheet_name)
        sfh = ws.cell(row=5, column=2).value
        mfh = ws.cell(row=5, column=3).value
        if sfh is not None and mfh is not None:
            records.append({
                "year": yr,
                "m2_per_person_sfh": float(sfh),
                "m2_per_person_mfh": float(mfh),
            })
    wb.close()

    return (
        pl.DataFrame(records)
        .sort("year")
        .with_columns(
            (pl.col("m2_per_person_sfh") * 0.30 + pl.col("m2_per_person_mfh") * 0.70)
            .round(1)
            .alias("m2_per_person_avg"),
        )
    )


# ────────────────────────────────────────────────────────────────────────
# BFS Buildings by Heating Source (2000 vs 2021)
# ────────────────────────────────────────────────────────────────────────

def download_sdmx_csv(
    directory: Path,
    url: str,
    filename: str,
    *,
    force: bool = False,
) -> Path:
    """Download CSV from a stats.swiss SDMX REST API endpoint, cache locally."""
    path = directory / filename
    if path.exists() and not force:
        print(f"  [cached] {filename}")
        return path
    print(f"  [downloading] {filename} ...")
    resp = httpx.get(
        url,
        headers={"Accept": "application/vnd.sdmx.data+csv;version=2.0.0"},
        follow_redirects=True,
        timeout=60.0,
    )
    resp.raise_for_status()
    path.write_bytes(resp.content)
    print(f"  [saved] {filename} ({len(resp.content) / 1024:.0f} KB)")
    return path


# ── GWR heating source code → label mapping ───────────────────────────
_GWR_HEATING_CODES: dict[str, str] = {
    "1": "Heat Pump",
    "2": "Gas",
    "3": "Heating Oil",
    "4": "Wood",
    "5": "Electricity",
    "6": "District Heating",
    "7": "Solar Thermal",
    "8": "Other",
    "9": "None",
}

_GWR_BUILDING_CODES: dict[str, str] = {
    "1021": "Single-Family (EFH)",
    "1025": "Multi-Family (MFH)",
    "1030": "Mixed-Use Residential",
    "1040": "Partial Residential",
}


def load_gwr_heating_by_type(csv_path: Path) -> pl.DataFrame:
    """Parse GWR SDMX CSV into a clean EFH/MFH × heating source table.

    Returns DataFrame with columns: ``year``, ``building_type``,
    ``heating_source``, ``building_count``.
    """
    df = pl.read_csv(csv_path, infer_schema_length=5000)

    # Cast to string first for reliable mapping (SDMX may infer as Int)
    return (
        df.with_columns(
            pl.col("GKATS").cast(pl.Utf8),
            pl.col("GWAERZH").cast(pl.Utf8),
        )
        .filter(pl.col("GKATS").is_in(["1021", "1025", "1030", "1040"]))
        .filter(~pl.col("GWAERZH").is_in(["_T", "9"]))  # drop totals & unheated
        .select(
            pl.col("TIME_PERIOD").cast(pl.Int64).alias("year"),
            pl.col("GKATS").replace_strict(
                _GWR_BUILDING_CODES, default="Other"
            ).alias("building_type"),
            pl.col("GWAERZH").replace_strict(
                _GWR_HEATING_CODES, default="Other"
            ).alias("heating_source"),
            pl.col("OBS_VALUE").cast(pl.Int64).alias("building_count"),
        )
        .sort("year", "building_type", "heating_source")
    )


def load_floor_area_by_period(xlsx_path: Path, *, year: int = 2024) -> pl.DataFrame:
    """Parse BFS floor area per person by construction period for EFH vs MFH.

    Returns DataFrame with columns: ``period``, ``m2_efh``, ``m2_mfh``.
    """
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[str(year)]

    records: list[dict] = []
    for row in range(6, ws.max_row + 1):
        label = ws.cell(row, 1).value
        efh = ws.cell(row, 2).value
        mfh = ws.cell(row, 3).value
        if label and efh is not None and mfh is not None:
            # Clean up German labels
            period = (
                str(label)
                .replace("Vor ", "< ")
                .replace("Zwischen ", "")
                .replace(" und ", "–")
                .replace(" erbaut", "")
                .replace("2)", "")
                .strip()
            )
            records.append({
                "period": period,
                "m2_efh": float(efh),
                "m2_mfh": float(mfh),
            })

    wb.close()
    return pl.DataFrame(records)



# ── BFS canton number ↔ abbreviation ────────────────────────────────
_CANTON_NUM_TO_ABBR: dict[int, str] = {
    1:"ZH", 2:"BE", 3:"LU", 4:"UR", 5:"SZ", 6:"OW", 7:"NW", 8:"GL",
    9:"ZG", 10:"FR", 11:"SO", 12:"BS", 13:"BL", 14:"SH", 15:"AR",
    16:"AI", 17:"SG", 18:"GR", 19:"AG", 20:"TG", 21:"TI", 22:"VD",
    23:"VS", 24:"NE", 25:"GE", 26:"JU",
}
_CANTON_ABBR_TO_NAME: dict[str, str] = {
    "ZH":"Zürich","BE":"Bern","LU":"Luzern","UR":"Uri","SZ":"Schwyz",
    "OW":"Obwalden","NW":"Nidwalden","GL":"Glarus","SG":"St. Gallen",
    "GR":"Graubünden","AG":"Aargau","TG":"Thurgau","TI":"Ticino",
    "VD":"Vaud","VS":"Valais","NE":"Neuchâtel","GE":"Genève",
    "JU":"Jura","FR":"Fribourg","SO":"Solothurn","BS":"Basel-Stadt",
    "BL":"Basel-Landschaft","SH":"Schaffhausen","AR":"Appenzell A.Rh.",
    "AI":"Appenzell I.Rh.","ZG":"Zug",
}


def topojson_to_geojson_cantons(topo_path: Path) -> dict:
    """Convert swiss-maps TopoJSON → GeoJSON FeatureCollection for cantons.

    Adds ``canton_id``, ``canton_abbr``, and ``canton_name`` properties
    so the GeoJSON can be matched to SDMX data via ``canton_abbr``.

    Uses a pure-Python arc decoding — no external ``topojson`` dependency.
    """
    topo = json.loads(topo_path.read_text())
    arcs_raw = topo["arcs"]
    transform = topo.get("transform")

    # Decode quantized arcs
    decoded_arcs: list[list[list[float]]] = []
    for arc in arcs_raw:
        coords: list[list[float]] = []
        x, y = 0.0, 0.0
        for pt in arc:
            x += pt[0]
            y += pt[1]
            if transform:
                coords.append([
                    x * transform["scale"][0] + transform["translate"][0],
                    y * transform["scale"][1] + transform["translate"][1],
                ])
            else:
                coords.append([x, y])
        decoded_arcs.append(coords)

    def _resolve_arc(idx: int) -> list[list[float]]:
        if idx >= 0:
            return decoded_arcs[idx]
        return decoded_arcs[~idx][::-1]

    def _ring(arc_indices: list[int]) -> list[list[float]]:
        ring: list[list[float]] = []
        for idx in arc_indices:
            pts = _resolve_arc(idx)
            ring.extend(pts if not ring else pts[1:])
        return ring

    features: list[dict] = []
    for geom in topo["objects"]["cantons"]["geometries"]:
        cid: int = geom["id"]
        abbr = _CANTON_NUM_TO_ABBR.get(cid, "??")
        name = _CANTON_ABBR_TO_NAME.get(abbr, "")
        props = {"canton_id": cid, "canton_abbr": abbr, "canton_name": name}

        if geom["type"] == "Polygon":
            coordinates = [_ring(ring) for ring in geom["arcs"]]
            geo_type = "Polygon"
        elif geom["type"] == "MultiPolygon":
            coordinates = [[_ring(ring) for ring in poly] for poly in geom["arcs"]]
            geo_type = "MultiPolygon"
        else:
            continue

        features.append({
            "type": "Feature",
            "id": cid,
            "properties": props,
            "geometry": {"type": geo_type, "coordinates": coordinates},
        })

    return {"type": "FeatureCollection", "features": features}


def load_gwr_canton_hp_share(csv_path: Path, *, year: int = 2024) -> pl.DataFrame:
    """Compute heat pump adoption share by canton from SDMX canton-level CSV.

    Returns DataFrame with columns: ``canton_id``, ``canton_abbr``,
    ``canton_name``, ``hp_share``, ``total_buildings``, ``hp_buildings``.
    """
    df = pl.read_csv(csv_path, infer_schema_length=10_000)

    # Filter: all building types (_T), specific year, exclude national total
    df = (
        df.with_columns(
            pl.col("GKATS").cast(pl.Utf8),
            pl.col("GWAERZH").cast(pl.Utf8),
            pl.col("KANTONSNUMMER").cast(pl.Utf8),
        )
        .filter(pl.col("GKATS") == "_T")          # all building types
        .filter(pl.col("TIME_PERIOD") == year)
        .filter(pl.col("KANTONSNUMMER") != "8100") # exclude national total
        .filter(~pl.col("GWAERZH").is_in(["_T"]))  # exclude totals but keep "9" (none)
    )

    # Total buildings per canton
    totals = (
        df.group_by("KANTONSNUMMER")
        .agg(pl.col("OBS_VALUE").sum().alias("total_buildings"))
    )

    # HP buildings per canton (GWAERZH == "1" = Heat Pump)
    hp = (
        df.filter(pl.col("GWAERZH") == "1")
        .group_by("KANTONSNUMMER")
        .agg(pl.col("OBS_VALUE").sum().alias("hp_buildings"))
    )

    abbr_to_id = {v: k for k, v in _CANTON_NUM_TO_ABBR.items()}

    merged = (
        totals.join(hp, on="KANTONSNUMMER", how="left")
        .with_columns(pl.col("hp_buildings").fill_null(0))
        .with_columns(
            (pl.col("hp_buildings") / pl.col("total_buildings") * 100)
            .round(1)
            .alias("hp_share"),
        )
        .with_columns(
            pl.col("KANTONSNUMMER").alias("canton_abbr"),
        )
        .with_columns(
            pl.col("canton_abbr")
            .map_elements(lambda a: abbr_to_id.get(a, 0), return_dtype=pl.Int64)
            .alias("canton_id"),
            pl.col("canton_abbr")
            .map_elements(lambda a: _CANTON_ABBR_TO_NAME.get(a, ""), return_dtype=pl.Utf8)
            .alias("canton_name"),
        )
        .select("canton_id", "canton_abbr", "canton_name",
                "hp_share", "total_buildings", "hp_buildings")
        .sort("canton_id")
    )
    return merged


def load_buildings_heating(xlsx_path: Path) -> pl.DataFrame:
    """Parse BFS buildings by heating energy source for 2000 and 2021.

    Returns DataFrame with columns: ``source``, ``buildings_2000``,
    ``buildings_2021``, ``change_pct``.
    """
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    # 2021
    ws21 = wb["2021"]
    heating_2021 = {
        "Heat Pump": ws21.cell(6, 3).value,
        "Gas": ws21.cell(6, 4).value,
        "Heating Oil": ws21.cell(6, 5).value,
        "Wood": ws21.cell(6, 6).value,
        "Electricity": ws21.cell(6, 7).value,
        "Solar Thermal": ws21.cell(6, 8).value,
        "District Heating": ws21.cell(6, 9).value,
        "Other": ws21.cell(6, 10).value,
    }

    # 2000
    ws00 = wb["2000"]
    heating_2000 = {
        "Heat Pump": ws00.cell(6, 3).value,
        "Gas": ws00.cell(6, 4).value,
        "Heating Oil": ws00.cell(6, 5).value,
        "Wood": ws00.cell(6, 7).value,
        "Electricity": ws00.cell(6, 8).value,
        "Solar Thermal": ws00.cell(6, 9).value,
        "District Heating": ws00.cell(6, 10).value,
        "Other": (
            int(ws00.cell(6, 6).value or 0) + int(ws00.cell(6, 11).value or 0)
        ),
    }

    wb.close()

    sources = list(heating_2021.keys())
    vals_2000 = [int(heating_2000.get(s, 0) or 0) for s in sources]
    vals_2021 = [int(heating_2021.get(s, 0) or 0) for s in sources]

    return pl.DataFrame({
        "source": sources,
        "buildings_2000": vals_2000,
        "buildings_2021": vals_2021,
    }).with_columns(
        (
            (pl.col("buildings_2021") - pl.col("buildings_2000"))
            / pl.col("buildings_2000").cast(pl.Float64)
            * 100
        )
        .round(1)
        .alias("change_pct"),
    )
